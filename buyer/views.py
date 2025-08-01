from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .models import Buyer, Cart, CartItem, Wishlist, Address, BuyerNotification
from .forms import BuyerUpdateForm
from seller.models import Product
from .utils import get_cart_from_cookie, set_cart_cookie, clear_cart_cookie
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .utils_order import group_cart_items_by_seller, calculate_order_summary, generate_order_id
from .models import Order, OrderItem
import base64
import stripe
import json
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from .models import Payment
from seller.models import GiftBoxCampaign, SellerGiftBoxParticipation, Seller, Product, Promotion
from .models import GiftBoxOrder
from django.urls import reverse
from django.db import transaction
from datetime import datetime, timedelta
import json
import csv
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
import os
import django
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import timedelta
import json
from .models import Buyer, Cart, CartItem, Wishlist, Order, OrderItem, Address, BuyerNotification, GiftBoxOrder
from seller.models import Seller, Product, ProductReview, SellerReview, Category, Brand
from .forms import BuyerRegistrationForm, BuyerUpdateForm, AddressForm

# Setup Django for potential external script usage
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'MarketVibe.settings')
django.setup()

from .models import Buyer, Order, GiftBoxOrder, Wishlist, Cart, BuyerNotification
from seller.models import Product, Seller


# Create your views here.
@login_required
def buyer_profile_view(request):
    try:
        buyer = Buyer.objects.get(email=request.user.email)
        
        # Handle form submission for profile update
        if request.method == 'POST':
            form = BuyerUpdateForm(request.POST, request.FILES, instance=buyer)
            if form.is_valid():
                buyer = form.save()
                messages.success(request, 'Buyer profile updated successfully.')
                return redirect('buyer:buyer_profile')
            else:
                messages.error(request, f'Please fix the errors below: {form.errors}')
        else:
            form = BuyerUpdateForm(instance=buyer)
        
        # Get relevant data for the profile page
        total_orders = Order.objects.filter(buyer=buyer).count()
        giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).count()
        total_all_orders = total_orders + giftbox_orders
        
        # Get wishlist count
        wishlist_count = Wishlist.objects.filter(buyer=buyer).count()
        
        # Get cart count
        try:
            cart = Cart.objects.get(buyer=buyer)
            cart_count = cart.total_quantity
        except Cart.DoesNotExist:
            cart_count = 0
        
        # Get recent orders
        recent_orders = Order.objects.filter(buyer=buyer).order_by('-created_at')[:5]
        recent_giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).order_by('-created_at')[:5]
        
        # Get notifications
        notifications = BuyerNotification.objects.filter(buyer=buyer, is_read=False).order_by('-created_at')[:5]
        
        context = {
            'buyer': buyer,
            'image_base64': buyer.get_image_base64(),
            'form': form,
            'total_orders': total_all_orders,
            'wishlist_count': wishlist_count,
            'cart_count': cart_count,
            'recent_orders': recent_orders,
            'recent_giftbox_orders': recent_giftbox_orders,
            'notifications': notifications,
        }
        return render(request, 'buyer/buyer_profile.html', context)
    except Buyer.DoesNotExist:
        messages.error(request, 'Buyer profile not found.')
        return redirect('sellers:login')


# UPDATE Buyer
@login_required
def update_buyer_view(request):
    try:
        buyer = Buyer.objects.get(email=request.user.email)
        if request.method == 'POST':
            form = BuyerUpdateForm(request.POST, request.FILES, instance=buyer)
            if form.is_valid():
                buyer = form.save()
                messages.success(request, 'Buyer profile updated successfully.')
                return redirect('buyer:buyer_profile')
        else:
            form = BuyerUpdateForm(instance=buyer)
        return render(request, 'buyer/buyer_update.html', {'form': form})
    except Buyer.DoesNotExist:
        messages.error(request, 'Buyer profile not found.')
        return redirect('sellers:login')


# DELETE Buyer
@login_required
def delete_buyer_view(request):
    try:
        buyer = Buyer.objects.get(email=request.user.email)
        if request.method == 'POST':
            # Delete the buyer profile
            buyer.delete()
            messages.success(request, 'Your buyer account has been deleted.')
            return redirect('sellers:login')
        return render(request, 'buyer/buyer_delete.html')
    except Buyer.DoesNotExist:
        messages.error(request, 'Buyer profile not found.')
        return redirect('sellers:login')


def get_cart(request):
    if request.user.is_authenticated:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        items = [
            {
                "product_id": item.product.id,
                "name": item.product.name,
                "price": item.product.final_price,
                "quantity": item.quantity,
                "total_price": item.get_total_price()
            }
            for item in cart.items.select_related('product')
        ]
        subtotal = sum(item["total_price"] for item in items)
        discount = 0
        tax = 0
        total = subtotal
        return {
            "items": items,
            "is_authenticated": True,
            "subtotal": subtotal,
            "discount": discount,
            "tax": tax,
            "total": total
        }
    else:
        cart_data = get_cart_from_cookie(request)
        # Calculate subtotal, discount, tax, total for guest cart
        items = []
        subtotal = 0
        for entry in cart_data.get('items', []):
            try:
                product = Product.objects.get(id=entry['product_id'])
                total_price = product.final_price * entry['quantity']
                items.append({
                    "product_id": product.id,
                    "name": product.name,
                    "price": product.final_price,
                    "quantity": entry['quantity'],
                    "total_price": total_price
                })
                subtotal += total_price
            except Product.DoesNotExist:
                continue
        discount = 0
        tax = 0
        total = subtotal
        cart_data.update({
            "items": items,
            "subtotal": subtotal,
            "discount": discount,
            "tax": tax,
            "total": total
        })
        return cart_data

@login_required(login_url='/seller/login/')
def add_to_cart(request):
    product_id = int(request.POST.get('product_id'))
    quantity = int(request.POST.get('quantity', 1))
    product = get_object_or_404(Product, id=product_id)
    if request.user.is_authenticated:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        cart.add_product(product, quantity)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        else:
            # Redirect to cart page after adding
            return redirect('buyer:cart_page')  # Make sure this is the correct name for your cart page view
    else:
        cart_data = get_cart_from_cookie(request)
        for item in cart_data["items"]:
            if item["product_id"] == product_id:
                item["quantity"] += quantity
                break
        else:
            cart_data["items"].append({"product_id": product_id, "quantity": quantity})
        response = JsonResponse({"success": True, "cart": cart_data})
        set_cart_cookie(response, cart_data)
        return response

@login_required
def remove_from_cart(request):
    product_id = int(request.POST.get('product_id'))
    if request.user.is_authenticated:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        product = get_object_or_404(Product, id=product_id)
        cart.remove_product(product)
        cart_data = get_cart(request)
        if isinstance(cart_data, JsonResponse):
            import json
            cart_data = json.loads(cart_data.content)
        return JsonResponse({"success": True, "cart": cart_data})
    else:
        cart_data = get_cart_from_cookie(request)
        new_items = [item for item in cart_data["items"] if item["product_id"] != product_id]
        cart_data["items"] = new_items
        response = JsonResponse({"success": True, "cart": cart_data})
        set_cart_cookie(response, cart_data)
        return response

@login_required
@require_POST
def update_cart_quantity(request):
    try:
        print("=== Starting cart quantity update ===")
        product_id = int(request.POST.get('product_id'))
        quantity = int(request.POST.get('quantity', 1))
        print(f"Received request - Product ID: {product_id}, Quantity: {quantity}")
        
        # Get the product first to validate it exists
        try:
            product = Product.objects.get(id=product_id)
            print(f"Found product: {product.name}, Price: {product.final_price:.2f}")
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Product not found'
            })

        if request.user.is_authenticated:
            try:
                buyer = Buyer.objects.get(email=request.user.email)
                cart, _ = Cart.objects.get_or_create(buyer=buyer)
                
                if quantity > 0:
                    cart_item = cart.update_quantity(product, quantity)
                else:
                    cart.remove_product(product)
                
                # Get updated cart data
                cart_data = cart.get_cart_data()
                print(f"Cart data after update: {cart_data}")
                return JsonResponse(cart_data)
                
            except Buyer.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Buyer not found'
                })
            except Exception as e:
                print(f"Error updating cart: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'error': f'Error updating cart: {str(e)}'
                })
        else:
            # Handle guest cart in session
            cart_data = request.session.get('cart', {'items': [], 'coupon_code': None})
            
            # Update or remove item
            updated_items = []
            subtotal = 0
            cart_count = 0
            
            for item in cart_data['items']:
                if item['product_id'] == product_id:
                    if quantity > 0:
                        item['quantity'] = quantity
                        updated_items.append(item)
                else:
                    updated_items.append(item)
                    
            cart_data['items'] = updated_items
            
            # Calculate totals
            for item in cart_data['items']:
                try:
                    prod = Product.objects.get(id=item['product_id'])
                    subtotal += round(prod.final_price * item['quantity'], 2)
                    cart_count += item['quantity']
                except Product.DoesNotExist:
                    continue
            
            # Apply discount
            discount_percentage = 0.20 if cart_data.get('coupon_code') in ["MarketVibe27", "Shopping24/7"] else 0.10
            discount = round(subtotal * discount_percentage, 2)
            tax = round((subtotal - discount) * 0.10, 2)
            total = round(subtotal - discount + tax, 2)
            
            # Update session
            request.session['cart'] = cart_data
            request.session.modified = True
            
            response_data = {
                'success': True,
                'cart_count': cart_count,
                'subtotal': float(subtotal),
                'discount_percentage': discount_percentage * 100,
                'discount': float(discount),
                'tax': float(tax),
                'total': float(total),
                'coupon_code': cart_data.get('coupon_code')
            }
            
            return JsonResponse(response_data)
            
    except Exception as e:
        print(f"Unexpected error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })
@login_required
@require_POST
def apply_coupon(request):
    try:
        code = request.POST.get('code')
        if not code:
            return JsonResponse({
                'success': False,
                'error': 'No coupon code provided'
            })
            
        if request.user.is_authenticated:
            buyer = Buyer.objects.get(email=request.user.email)
            cart, _ = Cart.objects.get_or_create(buyer=buyer)
            
            if cart.apply_coupon(code):
                return JsonResponse(cart.get_cart_data())
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid coupon code'
                })
        else:
            # Handle guest cart
            cart_data = request.session.get('cart', {'items': [], 'coupon_code': None})
            
            if code in ["MarketVibe27", "Shopping24/7"]:
                cart_data['coupon_code'] = code
                request.session['cart'] = cart_data
                request.session.modified = True
                
                # Recalculate totals
                subtotal = 0
                cart_count = 0
                
                for item in cart_data['items']:
                    try:
                        prod = Product.objects.get(id=item['product_id'])
                        subtotal += round(prod.final_price * item['quantity'], 2)
                        cart_count += item['quantity']
                    except Product.DoesNotExist:
                        continue
                
                discount = round(subtotal * 0.20, 2)  # 20% discount for valid coupon
                tax = round((subtotal - discount) * 0.10, 2)
                total = round(subtotal - discount + tax, 2)
                
                return JsonResponse({
                    'success': True,
                    'cart_count': cart_count,
                    'subtotal': float(subtotal),
                    'discount_percentage': 20.0,
                    'discount': float(discount),
                    'tax': float(tax),
                    'total': float(total),
                    'coupon_code': code
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid coupon code'
                })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def clear_cart(request):
    if request.user.is_authenticated:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        cart.clear()
        return get_cart(request)
    else:
        response = JsonResponse({"success": True, "cart": {"items": []}})
        clear_cart_cookie(response)
        return response

@login_required
@require_POST
def add_to_wishlist(request):
    print("=== Adding to wishlist ===")
    try:
        product_id = int(request.POST.get('product_id'))
        print(f"Product ID: {product_id}")
        
        if not request.user.is_authenticated:
            print("User not authenticated")
            return JsonResponse({
                'success': False,
                'error': 'Please login to save items'
            })
        
        try:
            product = Product.objects.get(id=product_id)
            print(f"Found product: {product.name}")
        except Product.DoesNotExist:
            print(f"Product not found with ID: {product_id}")
            return JsonResponse({
                'success': False,
                'error': 'Product not found'
            })
        
        buyer = get_object_or_404(Buyer, email=request.user.email)
        print(f"Found buyer: {buyer.name}")
        
        # Check if this is a request from cart page (allow moving from cart to wishlist)
        is_from_cart = request.POST.get('from_cart', 'false').lower() == 'true'
        
        # Only check if item is in cart if not coming from cart page
        if not is_from_cart:
            cart = Cart.objects.filter(buyer=buyer).first()
            if cart and cart.items.filter(product=product).exists():
                return JsonResponse({
                    'success': False,
                    'error': 'This item is already in your cart.'
                })
        
        # Check if already in wishlist
        wishlist_item, created = Wishlist.objects.get_or_create(
            buyer=buyer,
            product=product
        )
        print(f"Wishlist item {'created' if created else 'already exists'}")
        
        return JsonResponse({
            'success': True,
            'message': 'Added to wishlist',
            'in_wishlist': True
        })
        
    except Exception as e:
        print(f"Error adding to wishlist: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_POST
def check_cart_status(request):
    print("=== Checking cart status ===")
    try:
        product_id = int(request.POST.get('product_id'))
        print(f"Product ID: {product_id}")
        
        if not request.user.is_authenticated:
            print("User not authenticated")
            return JsonResponse({
                'success': False,
                'error': 'Please login to check cart'
            })
            
        buyer = get_object_or_404(Buyer, email=request.user.email)
        print(f"Found buyer: {buyer.name}")
        
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        in_cart = cart.items.filter(product_id=product_id).exists()
        print(f"Product in cart: {in_cart}")
        
        return JsonResponse({
            'success': True,
            'in_cart': in_cart
        })
        
    except Exception as e:
        print(f"Error checking cart status: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_POST
def check_wishlist_status(request):
    print("=== Checking wishlist status ===")
    try:
        product_id = int(request.POST.get('product_id'))
        print(f"Product ID: {product_id}")
        
        if not request.user.is_authenticated:
            print("User not authenticated")
            return JsonResponse({
                'success': False,
                'error': 'Please login to check wishlist'
            })
            
        buyer = get_object_or_404(Buyer, email=request.user.email)
        print(f"Found buyer: {buyer.name}")
        
        in_wishlist = Wishlist.objects.filter(buyer=buyer, product_id=product_id).exists()
        print(f"Product in wishlist: {in_wishlist}")
        
        return JsonResponse({
            'success': True,
            'in_wishlist': in_wishlist
        })
        
    except Exception as e:
        print(f"Error checking wishlist status: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_POST
def remove_from_wishlist(request):
    print("=== Removing from wishlist ===")
    try:
        product_id = int(request.POST.get('product_id'))
        print(f"Product ID: {product_id}")
        
        if not request.user.is_authenticated:
            print("User not authenticated")
            return JsonResponse({
                'success': False,
                'error': 'Please login to manage wishlist'
            })
            
        buyer = get_object_or_404(Buyer, email=request.user.email)
        print(f"Found buyer: {buyer.name}")
        
        result = Wishlist.objects.filter(buyer=buyer, product_id=product_id).delete()
        print(f"Delete result: {result}")
        
        return JsonResponse({
            'success': True,
            'message': 'Removed from wishlist'
        })
        
    except Exception as e:
        print(f"Error removing from wishlist: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_POST
def remove_all_from_cart(request):
    try:
        if request.user.is_authenticated:
            buyer = get_object_or_404(Buyer, email=request.user.email)
            cart, _ = Cart.objects.get_or_create(buyer=buyer)
            cart.clear()  # Using the existing clear method from Cart model
            
            return JsonResponse({
                'success': True,
                'message': 'All items removed from cart',
                'cart_count': 0,
                'subtotal': 0,
                'tax': 0,
                'discount': 0,
                'total': 0
            })
        else:
            # Handle guest cart
            request.session['cart'] = {'items': []}
            request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'All items removed from cart',
                'cart_count': 0,
                'subtotal': 0,
                'tax': 0,
                'discount': 0,
                'total': 0
            })
            
    except Exception as e:
        print(f"Error removing all items from cart: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

# Update the cart_page_view to include wishlist items
def cart_page_view(request):
    if request.user.is_authenticated:
        buyer = Buyer.objects.get(email=request.user.email)
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        items = cart.items.select_related('product')
        cart_items = [{
            'product': item.product,
            'quantity': item.quantity,
            'total_price': item.get_total_price()
        } for item in items]
        
        # Get wishlist items
        wishlist_items = Wishlist.objects.filter(buyer=buyer).select_related('product')
        
        subtotal = cart.subtotal or 0
        discount = cart.discount_amount or 0
        tax = cart.tax or 0
        total = cart.total or 0
        discount_percentage = (cart.get_discount_percentage() or 0) * 100
    else:
        cart_data = request.session.get('cart', {'items': []})
        cart_items = []
        wishlist_items = []  # Empty list for non-authenticated users
        subtotal = 0
        
        for entry in cart_data['items']:
            try:
                product = Product.objects.get(id=entry['product_id'])
                # Use calculated final price with fallback
                price = getattr(product, 'calculated_final_price', 0) or 0
                item_total = price * entry['quantity']
                cart_items.append({
                    'product': product,
                    'quantity': entry['quantity'],
                    'total_price': item_total
                })
                subtotal += item_total
            except Product.DoesNotExist:
                continue
                
        discount_percentage = 0.20 if cart_data.get('coupon_code') in ["MarketVibe27", "Shopping24/7"] else 0.10
        discount = round(subtotal * discount_percentage, 2)
        tax = round((subtotal - discount) * 0.10, 2)
        total = round(subtotal - discount + tax, 2)
        discount_percentage *= 100

    context = {
        'cart_items': cart_items,
        'wishlist_items': wishlist_items,
        'subtotal': subtotal,
        'discount': discount,
        'tax': tax,
        'total': total,
        'discount_percentage': discount_percentage
    }
    return render(request, 'buyer/cart_page.html', context)

@login_required
def checkout_view(request):
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Check if this is a gift box checkout
    giftbox_data = request.session.get('giftbox_data')
    is_giftbox_checkout = giftbox_data is not None
    
    # Check if this is a promotion checkout
    promotion_data = request.session.get('promotion_data')
    is_promotion_checkout = promotion_data is not None
    
    # Clear conflicting session data to ensure proper checkout type
    if is_giftbox_checkout and is_promotion_checkout:
        # If both are present, we need to determine which one is more recent
        # Check if we came from a promotion checkout URL
        referer = request.META.get('HTTP_REFERER', '')
        if 'promotion/checkout' in referer or 'promotion/' in referer:
            # Clear giftbox data if we came from promotion checkout
            if 'giftbox_data' in request.session:
                del request.session['giftbox_data']
            is_giftbox_checkout = False
        else:
            # Clear promotion data if we came from giftbox checkout
            if 'promotion_data' in request.session:
                del request.session['promotion_data']
            is_promotion_checkout = False
    
    # Add debug logging
    print(f"DEBUG: Checkout view - Giftbox: {is_giftbox_checkout}, Promotion: {is_promotion_checkout}")
    print(f"DEBUG: Session data - Giftbox: {giftbox_data}, Promotion: {promotion_data}")
    print(f"DEBUG: Referer: {request.META.get('HTTP_REFERER', '')}")
    
    # If neither giftbox nor promotion checkout, this is a regular cart checkout
    if not is_giftbox_checkout and not is_promotion_checkout:
        # Clear any existing session data for other checkout types
        if 'giftbox_data' in request.session:
            del request.session['giftbox_data']
        if 'promotion_data' in request.session:
            del request.session['promotion_data']
    else:
        # If we have session data but the referer indicates we came from cart page,
        # we should clear the session data and treat as regular checkout
        referer = request.META.get('HTTP_REFERER', '')
        if 'cart/page/' in referer or 'cart/' in referer:
            print(f"DEBUG: Clearing session data because we came from cart page")
            if 'giftbox_data' in request.session:
                del request.session['giftbox_data']
            if 'promotion_data' in request.session:
                del request.session['promotion_data']
            is_giftbox_checkout = False
            is_promotion_checkout = False
    
    if is_giftbox_checkout:
        # Clear any existing promotion data when doing giftbox checkout
        if 'promotion_data' in request.session:
            del request.session['promotion_data']
        
        # Handle gift box checkout
        seller_id = giftbox_data.get('seller_id')
        campaign_id = giftbox_data.get('campaign_id')
        buyer_message = giftbox_data.get('buyer_message', '')
        giftbox_price = giftbox_data.get('price', 0)
        
        # Get seller object
        try:
            seller = get_object_or_404(Seller, id=seller_id)
        except:
            # If seller doesn't exist, clear the session and redirect to cart
            print(f"DEBUG: Seller {seller_id} not found, clearing session")
            if 'giftbox_data' in request.session:
                del request.session['giftbox_data']
            messages.error(request, 'The seller for this gift box is no longer available.')
            return redirect('buyer:cart_page')
        campaign = get_object_or_404(GiftBoxCampaign, id=campaign_id) if campaign_id else None
        
        # Create gift box seller group
        grouped_sellers = [{
            'seller': {'shop_name': seller.shop_name},
            'products': [{
                'name': f'Gift Box - {campaign.name if campaign else "Surprise Box"}',
                'image_url': '',  # No image for gift box
                'attributes': f'Message: {buyer_message}' if buyer_message else 'Surprise contents',
                'quantity': 1,
                'price': giftbox_price,
                'is_giftbox': True,
            }],
            'subtotal': giftbox_price,
        }]
        
        # Calculate order summary for gift box
        subtotal = giftbox_price
        shipping = 15.99
        tax = round(subtotal * 0.125, 2)  # 12.5% tax
        total = subtotal + shipping + tax
        total_items = 1
        
        order_summary = {
            'subtotal': subtotal,
            'shipping': shipping,
            'tax': tax,
            'total': total,
            'total_items': total_items,
        }
        
        # Get buyer's address (it's a TextField, not an object)
        buyer_address = getattr(buyer, 'address', '')
        shipping_address = {
            'street': buyer_address,
            'city': '',
            'zip_code': '',
            'country': 'US',
        }
        
        context = {
            'grouped_sellers': grouped_sellers,
            'order_summary': order_summary,
            'shipping_address': shipping_address,
            'buyer': buyer,
            'cart': None,  # No cart for gift box
            'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY,
            'is_giftbox_checkout': True,
            'is_promotion_checkout': False,
            'is_regular_checkout': False,
            'giftbox_data': giftbox_data,
        }
    elif is_promotion_checkout:
        # Clear any existing giftbox data when doing promotion checkout
        if 'giftbox_data' in request.session:
            del request.session['giftbox_data']
        
        # Handle promotion checkout
        promotion_id = promotion_data.get('promotion_id')
        promotion_name = promotion_data.get('promotion_name', 'Promotion')
        promotion_type = promotion_data.get('promotion_type', 'percentage')
        discount_value = promotion_data.get('discount_value', 0)
        total_amount = promotion_data.get('total_amount', 0)
        seller_id = promotion_data.get('seller_id')
        seller_name = promotion_data.get('seller_name', 'Seller')
        
        # Get seller object
        try:
            seller = get_object_or_404(Seller, id=seller_id)
        except:
            # If seller doesn't exist, clear the session and redirect to cart
            print(f"DEBUG: Seller {seller_id} not found, clearing session")
            if 'promotion_data' in request.session:
                del request.session['promotion_data']
            messages.error(request, 'The seller for this promotion is no longer available.')
            return redirect('buyer:cart_page')
        
        # Get the actual promotion object to access its products
        try:
            promotion_obj = Promotion.objects.get(id=promotion_id)
        except Promotion.DoesNotExist:
            promotion_obj = None
        
        # Create promotion seller group
        grouped_sellers = [{
            'seller': {'shop_name': seller_name},
            'promotion': promotion_obj,  # Include the promotion object
            'products': [{
                'name': f'Promotion: {promotion_name}',
                'image_url': '',  # No image for promotion
                'attributes': f'Discount: {discount_value}%' if promotion_type == 'percentage' else f'Discount: ${discount_value}',
                'quantity': 1,
                'price': total_amount,
                'is_promotion': True,
            }],
            'subtotal': total_amount,
        }]
        
        # Calculate order summary for promotion
        subtotal = total_amount
        shipping = 15.99
        tax = round(subtotal * 0.125, 2)  # 12.5% tax
        total = subtotal + shipping + tax
        total_items = 1
        
        order_summary = {
            'subtotal': subtotal,
            'shipping': shipping,
            'tax': tax,
            'total': total,
            'total_items': total_items,
        }
        
        # Get buyer's address (it's a TextField, not an object)
        buyer_address = getattr(buyer, 'address', '')
        shipping_address = {
            'street': buyer_address,
            'city': '',
            'zip_code': '',
            'country': 'US',
        }
        
        context = {
            'grouped_sellers': grouped_sellers,
            'order_summary': order_summary,
            'shipping_address': shipping_address,
            'buyer': buyer,
            'cart': None,  # No cart for promotion
            'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY,
            'is_giftbox_checkout': False,
            'is_promotion_checkout': True,
            'is_regular_checkout': False,
            'promotion_data': promotion_data,
        }
    else:
        # Handle regular cart checkout
        print(f"DEBUG: Processing regular cart checkout")
        cart, _ = Cart.objects.get_or_create(buyer=buyer)
        cart_items = cart.items.select_related('product__seller')
        
        # Check if cart has items
        if not cart_items.exists():
            print(f"DEBUG: Cart is empty, redirecting to cart page")
            messages.error(request, 'Your cart is empty. Please add items before checkout.')
            return redirect('buyer:cart_page')
        
        print(f"DEBUG: Cart has {cart_items.count()} items")

        grouped_sellers = []
        seller_map = {}
        print(f"DEBUG: Processing {cart_items.count()} cart items")
        for item in cart_items:
            seller = item.product.seller
            print(f"DEBUG: Processing item: {item.product.name} from seller: {seller.shop_name}")
            if seller.id not in seller_map:
                seller_map[seller.id] = {
                    'seller': {'shop_name': seller.shop_name},
                    'products': [],
                    'subtotal': 0,
                }
            img_obj = item.product.images.first()
            if img_obj and img_obj.image:
                image_url = f"data:image/jpeg;base64,{base64.b64encode(img_obj.image).decode()}"
            else:
                image_url = ""
            seller_map[seller.id]['products'].append({
                'name': item.product.name,
                'image_url': image_url,
                'attributes': ', '.join([f"{av.attribute.name}: {av.value}" for av in item.product.attribute_values.all()]),
                'quantity': item.quantity,
                'price': item.get_total_price() / item.quantity,  # Use the actual price per unit after discounts
            })
            seller_map[seller.id]['subtotal'] += item.get_total_price()  # Use cart item's total price
        grouped_sellers = list(seller_map.values())
        print(f"DEBUG: Created {len(grouped_sellers)} seller groups")

        # Use cart's calculated values instead of manual calculation
        subtotal = cart.subtotal
        shipping = 15.99  # Example fixed shipping
        tax = cart.tax  # Use cart's calculated tax
        total = cart.total + shipping  # Add shipping to cart's total
        total_items = cart.total_quantity  # Use cart's total quantity
        
        order_summary = {
            'subtotal': subtotal, 
            'shipping': shipping, 
            'tax': tax, 
            'total': total, 
            'total_items': total_items,
        }

        # Get buyer's address (it's a TextField, not an object)
        buyer_address = getattr(buyer, 'address', '')
        shipping_address = {
            'street': buyer_address,
            'city': '',
            'zip_code': '',
            'country': 'US',
        }

        context = {
            'grouped_sellers': grouped_sellers, 
            'order_summary': order_summary,
            'shipping_address': shipping_address, 
            'buyer': buyer,
            'cart': cart,  # Add cart to context for discount display
            'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY,
            'is_giftbox_checkout': False,
            'is_promotion_checkout': False,
            'is_regular_checkout': True,
        }
    
    return render(request, 'buyer/checkout_page.html', context)

@login_required
def order_summary_view(request):
    buyer = get_object_or_404(Buyer, email=request.user.email)
    cart, _ = Cart.objects.get_or_create(buyer=buyer)
    items = cart.items.select_related('product__seller')

    # Group items by seller
    sellers = {}
    for item in items:
        seller = item.product.seller
        if seller.id not in sellers:
            sellers[seller.id] = {
                'seller': seller,
                'products': []
            }
        sellers[seller.id]['products'].append(item)

    context = {
        'sellers': sellers.values()
    }
    return render(request, 'buyer/checkout_page.html', context)

@login_required
@require_POST
def place_order_view(request):
    """Handle COD order placement"""
    try:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        
        # Check if this is a gift box order
        giftbox_data = request.session.get('giftbox_data')
        is_giftbox_order = giftbox_data is not None
        
        # Check if this is a promotion order
        promotion_data = request.session.get('promotion_data')
        is_promotion_order = promotion_data is not None
        
        if is_giftbox_order:
            # Handle gift box order
            seller_id = giftbox_data.get('seller_id')
            campaign_id = giftbox_data.get('campaign_id')
            buyer_message = giftbox_data.get('buyer_message', '')
            
            seller = get_object_or_404(Seller, id=seller_id)
            campaign = get_object_or_404(GiftBoxCampaign, id=campaign_id) if campaign_id else None
            
            # Get shipping address from request
            import json
            data = json.loads(request.body)
            shipping_address = data.get('shipping_address', {})
            shipping_address_str = f"{shipping_address.get('street', '')}, {shipping_address.get('city', '')}, {shipping_address.get('zip_code', '')}, {shipping_address.get('country', '')}"
            
            # Create gift box order
            giftbox_order = GiftBoxOrder.objects.create(
                buyer=buyer,
                seller=seller,
                campaign=campaign,
                status='pending',
                buyer_message=buyer_message,
                delivery_address=shipping_address_str,
            )
            
            # Clear gift box data from session
            if 'giftbox_data' in request.session:
                del request.session['giftbox_data']
            
            return JsonResponse({
                'success': True,
                'orders': [giftbox_order.id],
                'is_giftbox': True
            })
        elif is_promotion_order:
            # Handle promotion order
            print(f"DEBUG: Processing promotion order")
            promotion_id = promotion_data.get('promotion_id')
            promotion_name = promotion_data.get('promotion_name', 'Promotion')
            total_amount = promotion_data.get('total_amount', 0)
            seller_id = promotion_data.get('seller_id')
            print(f"DEBUG: Promotion ID: {promotion_id}, Name: {promotion_name}, Amount: {total_amount}, Seller ID: {seller_id}")
            
            seller = get_object_or_404(Seller, id=seller_id)
            from seller.models import Promotion
            promotion = get_object_or_404(Promotion, id=promotion_id)
            
            # Get shipping address from request
            import json
            data = json.loads(request.body)
            shipping_address = data.get('shipping_address', {})
            shipping_address_str = f"{shipping_address.get('street', '')}, {shipping_address.get('city', '')}, {shipping_address.get('zip_code', '')}, {shipping_address.get('country', '')}"
            
            # Create promotion order (using regular Order model)
            print(f"DEBUG: Creating promotion order for promotion {promotion_id}")
            order = Order.objects.create(
                buyer=buyer,
                seller=seller,
                status='pending',
                total_amount=total_amount,
                payment_status='pending',
                order_type='cod',
                delivery_address=shipping_address_str,
                notes=f'Promotion: {promotion_name}',
                promotion=promotion
            )
            print(f"DEBUG: Created promotion order ID: {order.id}")
            
            # Create order item for the promotion
            # Use the first product from the seller as a placeholder
            placeholder_product = seller.products.first()
            if not placeholder_product:
                # If no products exist, create a minimal order item
                return JsonResponse({'success': False, 'error': 'Seller has no products available'})
            
            OrderItem.objects.create(
                order=order,
                product=placeholder_product,
                price_at_purchase=total_amount,
                quantity=1
            )
            
            # Increment promotion used_count
            try:
                promotion.used_count += 1
                promotion.save()
                print(f"DEBUG: Incremented promotion {promotion.id} used_count to {promotion.used_count}")
            except Exception as e:
                print(f"DEBUG: Error incrementing promotion used_count: {str(e)}")
            
            # Clear promotion data from session
            if 'promotion_data' in request.session:
                del request.session['promotion_data']
            
            return JsonResponse({
                'success': True,
                'orders': [order.id],
                'is_promotion': True
            })
        else:
            # Handle regular cart order
            cart, _ = Cart.objects.get_or_create(buyer=buyer)
            cart_items = cart.items.select_related('product__seller')
            
            if not cart_items.exists():
                return JsonResponse({'success': False, 'error': 'Cart is empty'})
            
            # Get shipping address from request
            import json
            data = json.loads(request.body)
            shipping_address = data.get('shipping_address', {})
            shipping_address_str = f"{shipping_address.get('street', '')}, {shipping_address.get('city', '')}, {shipping_address.get('zip_code', '')}, {shipping_address.get('country', '')}"
            
            # Create orders for each seller
            orders_created = []
            
            for item in cart_items:
                seller = item.product.seller
                
                # Create order
                order = Order.objects.create(
                    buyer=buyer,
                    seller=seller,
                    status='pending',
                    total_amount=item.get_total_price(),
                    payment_status='pending',
                    order_type='cod',
                    delivery_address=shipping_address_str,
                )
                
                # Create order item
                OrderItem.objects.create(
                    order=order,
                    product=item.product,
                    price_at_purchase=item.get_total_price() / item.quantity,  # Use the actual price per unit after discounts
                    quantity=item.quantity,
                )
                
                # Increment product order count and decrement stock
                item.product.order_count += 1
                item.product.stock -= item.quantity
                item.product.save()
                
                orders_created.append(order)
            
            # Clear cart after creating orders
            cart.clear()
            
            return JsonResponse({
                'success': True,
                'orders': [order.id for order in orders_created],
                'is_giftbox': False
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

stripe.api_key = settings.STRIPE_SECRET_KEY

def create_stripe_payment_intent(order, seller):
    # Calculate amounts
    amount = int(order.total_amount * 100)  # in cents
    commission = int(order.total_amount * 0.10 * 100)  # 10% commission, in cents

    payment_intent = stripe.PaymentIntent.create(
        amount=amount,
        currency="usd",
        payment_method_types=["card"],
        application_fee_amount=commission,
        transfer_data={
            "destination": seller.stripe_account_id,
        },
        metadata={
            "order_id": order.id,
        }
    )
    return payment_intent.client_secret

@login_required
@require_POST
def process_stripe_payment(request):
    """Handle Stripe payment processing during checkout"""
    try:
        buyer = get_object_or_404(Buyer, email=request.user.email)
        
        # Check if this is a gift box order
        giftbox_data = request.session.get('giftbox_data')
        is_giftbox_order = giftbox_data is not None
        
        # Check if this is a promotion order
        promotion_data = request.session.get('promotion_data')
        is_promotion_order = promotion_data is not None
        
        if is_giftbox_order:
            # Handle gift box Stripe payment
            seller_id = giftbox_data.get('seller_id')
            campaign_id = giftbox_data.get('campaign_id')
            buyer_message = giftbox_data.get('buyer_message', '')
            giftbox_price = giftbox_data.get('price', 0)
            
            seller = get_object_or_404(Seller, id=seller_id)
            campaign = get_object_or_404(GiftBoxCampaign, id=campaign_id) if campaign_id else None
            
            # Check if seller has Stripe account
            if not seller.stripe_account_id:
                return JsonResponse({
                    'success': False,
                    'error': f'Seller {seller.shop_name} is not set up for online payments. Please contact the seller or choose Cash on Delivery.'
                })
            
            # Get shipping address from form
            shipping_address = request.POST.get('shipping_address', '')
            
            # Calculate total with shipping and tax
            subtotal = giftbox_price
            shipping = 15.99
            tax = round(subtotal * 0.125, 2)  # 12.5% tax
            total_amount = subtotal + shipping + tax
            
            # Create gift box order
            giftbox_order = GiftBoxOrder.objects.create(
                buyer=buyer,
                seller=seller,
                campaign=campaign,
                status='pending',
                buyer_message=buyer_message,
                delivery_address=shipping_address,
            )
            
            # Create Stripe PaymentIntent for gift box
            amount = int(total_amount * 100)  # in cents
            commission = int(total_amount * 0.10 * 100)  # 10% commission, in cents
            
            payment_intent = stripe.PaymentIntent.create(
                amount=amount,
                currency="usd",
                payment_method_types=["card"],
                application_fee_amount=commission,
                transfer_data={
                    "destination": seller.stripe_account_id,
                },
                metadata={
                    "order_id": giftbox_order.id,
                    "order_type": "giftbox",
                }
            )
            
            # Clear gift box data from session
            if 'giftbox_data' in request.session:
                del request.session['giftbox_data']
            
            return JsonResponse({
                'success': True,
                'payment_intents': [{
                    'order_id': giftbox_order.id,
                    'client_secret': payment_intent.client_secret,
                    'amount': total_amount
                }],
                'orders': [giftbox_order.id],
                'is_giftbox': True
            })
        elif is_promotion_order:
            # Handle promotion Stripe payment
            print(f"DEBUG: Processing promotion Stripe payment")
            promotion_id = promotion_data.get('promotion_id')
            promotion_name = promotion_data.get('promotion_name', 'Promotion')
            total_amount = promotion_data.get('total_amount', 0)
            seller_id = promotion_data.get('seller_id')
            print(f"DEBUG: Promotion ID: {promotion_id}, Name: {promotion_name}, Amount: {total_amount}, Seller ID: {seller_id}")
            
            seller = get_object_or_404(Seller, id=seller_id)
            from seller.models import Promotion
            promotion = get_object_or_404(Promotion, id=promotion_id)
            
            # Check if seller has Stripe account
            if not seller.stripe_account_id:
                return JsonResponse({
                    'success': False,
                    'error': f'Seller {seller.shop_name} is not set up for online payments. Please contact the seller or choose Cash on Delivery.'
                })
            
            # Get shipping address from form and format it
            shipping_address_raw = request.POST.get('shipping_address', '')
            # Parse the shipping address if it's JSON, otherwise use as is
            import json
            try:
                shipping_address_data = json.loads(shipping_address_raw) if shipping_address_raw else {}
                shipping_address_str = f"{shipping_address_data.get('street', '')}, {shipping_address_data.get('city', '')}, {shipping_address_data.get('zip_code', '')}, {shipping_address_data.get('country', '')}"
            except (json.JSONDecodeError, TypeError):
                shipping_address_str = shipping_address_raw
            
            # Calculate total with shipping and tax
            subtotal = total_amount
            shipping = 15.99
            tax = round(subtotal * 0.125, 2)  # 12.5% tax
            final_total = subtotal + shipping + tax
            
            # Create promotion order (using regular Order model)
            print(f"DEBUG: Creating promotion order for promotion {promotion_id}")
            order = Order.objects.create(
                buyer=buyer,
                seller=seller,
                status='pending',
                total_amount=final_total,
                payment_status='pending',
                order_type='stripe',
                delivery_address=shipping_address_str,
                notes=f'Promotion: {promotion_name}',
                promotion=promotion
            )
            print(f"DEBUG: Created promotion order ID: {order.id}")
            
            # Create order item for the promotion
            # Use the first product from the seller as a placeholder
            placeholder_product = seller.products.first()
            if not placeholder_product:
                # If no products exist, create a minimal order item
                return JsonResponse({'success': False, 'error': 'Seller has no products available'})
            
            OrderItem.objects.create(
                order=order,
                product=placeholder_product,
                price_at_purchase=final_total,
                quantity=1
            )
            
            # Increment promotion used_count
            try:
                promotion.used_count += 1
                promotion.save()
                print(f"DEBUG: Incremented promotion {promotion.id} used_count to {promotion.used_count}")
            except Exception as e:
                print(f"DEBUG: Error incrementing promotion used_count: {str(e)}")
            
            # Create Stripe PaymentIntent for promotion
            amount = int(final_total * 100)  # in cents
            commission = int(final_total * 0.10 * 100)  # 10% commission, in cents
            
            payment_intent = stripe.PaymentIntent.create(
                amount=amount,
                currency="usd",
                payment_method_types=["card"],
                application_fee_amount=commission,
                transfer_data={
                    "destination": seller.stripe_account_id,
                },
                metadata={
                    "order_id": order.id,
                    "order_type": "promotion",
                }
            )
            
            # Clear promotion data from session
            if 'promotion_data' in request.session:
                del request.session['promotion_data']
            
            return JsonResponse({
                'success': True,
                'payment_intents': [{
                    'order_id': order.id,
                    'client_secret': payment_intent.client_secret,
                    'amount': final_total
                }],
                'orders': [order.id],
                'is_promotion': True
            })
        else:
            # Handle regular cart Stripe payment
            cart, _ = Cart.objects.get_or_create(buyer=buyer)
            cart_items = cart.items.select_related('product__seller')
            
            if not cart_items.exists():
                return JsonResponse({'success': False, 'error': 'Cart is empty'})
            
            # Get shipping address from form
            shipping_address = request.POST.get('shipping_address', '')
            
            # Group items by seller and check Stripe setup
            seller_groups = {}
            stripe_ready_sellers = []
            non_stripe_sellers = []
            
            for item in cart_items:
                seller = item.product.seller
                if seller.id not in seller_groups:
                    seller_groups[seller.id] = {
                        'seller': seller,
                        'items': [],
                        'total': 0
                    }
                
                seller_groups[seller.id]['items'].append(item)
                seller_groups[seller.id]['total'] += item.get_total_price()
                
                # Check if seller has Stripe account
                if seller.stripe_account_id:
                    stripe_ready_sellers.append(seller.id)
                else:
                    non_stripe_sellers.append(seller.shop_name)
            
            # If any seller doesn't have Stripe, provide options
            if non_stripe_sellers:
                return JsonResponse({
                    'success': False,
                    'error': 'Some sellers are not set up for online payments',
                    'non_stripe_sellers': non_stripe_sellers,
                    'stripe_ready_sellers': stripe_ready_sellers,
                    'suggestion': 'Please contact these sellers to set up payments or choose Cash on Delivery'
                })
            
            # All sellers have Stripe - proceed with payment
            payment_intents = []
            orders_created = []
            
            for seller_id, group in seller_groups.items():
                seller = group['seller']
                
                # Calculate shipping per seller
                shipping_per_seller = 15.99 / len(seller_groups)

                # Create order
                order = Order.objects.create(
                    buyer=buyer,
                    seller=seller,
                    status='pending',
                    total_amount=group['total'] + shipping_per_seller,  # Add shipping here
                    payment_status='pending',
                    order_type='stripe',
                    delivery_address=shipping_address,
                )
                
                # Create order items
                for item in group['items']:
                    OrderItem.objects.create(
                        order=order,
                        product=item.product,
                        price_at_purchase=item.get_total_price() / item.quantity,  # Use the actual price per unit after discounts
                        quantity=item.quantity,
                    )
                    
                    # Increment product order count and decrement stock
                    item.product.order_count += 1
                    item.product.stock -= item.quantity
                    item.product.save()
                
                # Create Stripe PaymentIntent
                client_secret = create_stripe_payment_intent(order, seller)
                payment_intents.append({
                    'order_id': order.id,
                    'client_secret': client_secret,
                    'amount': order.total_amount
                })
                orders_created.append(order)
            
            # Clear cart after creating orders
            cart.clear()
            
            return JsonResponse({
                'success': True,
                'payment_intents': payment_intents,
                'orders': [order.id for order in orders_created],
                'is_giftbox': False
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
def stripe_webhook(request):
    import stripe
    from django.conf import settings
    from .models import Payment

    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    
    # For local development, if signature is missing, try to process without verification
    if not sig_header and settings.DEBUG:
        try:
            event = json.loads(payload)
        except json.JSONDecodeError:
            return HttpResponse(status=400)
    else:
        if not sig_header:
            return HttpResponse(status=400)
        
        try:
            event = stripe.Webhook.construct_event(
                payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
            )
        except ValueError:
            return HttpResponse(status=400)
        except stripe.error.SignatureVerificationError:
            return HttpResponse(status=400)

    if event['type'] == 'payment_intent.succeeded':
        intent = event['data']['object']
        order_id = intent['metadata'].get('order_id')
        
        if not order_id:
            print(f"DEBUG: No order_id found in metadata: {intent['metadata']}")
            return HttpResponse(status=400)
        
        try:
            order = Order.objects.get(id=order_id)
            order.payment_status = 'paid'
            order.status = 'processing'
            order.save()
            
            # Create payment record
            payment = Payment.objects.create(
                order=order,
                payment_method='stripe',
                transaction_id=intent.id,
                status='paid',
                payment_time=timezone.now()
            )
            print(f"DEBUG: Payment record created: {payment.id} for order {order_id}")
            
        except Order.DoesNotExist:
            print(f"DEBUG: Order {order_id} not found")
            return HttpResponse(status=404)
        except Exception as e:
            print(f"DEBUG: Error creating payment: {str(e)}")
            return HttpResponse(status=500)

    elif event['type'] == 'payment_intent.payment_failed':
        intent = event['data']['object']
        order_id = intent['metadata']['order_id']
        
        try:
            order = Order.objects.get(id=order_id)
            order.payment_status = 'failed'
            order.save()
            
            # Create payment record
            Payment.objects.create(
                order=order,
                payment_method='stripe',
                transaction_id=intent.id,
                status='failed',
                payment_time=timezone.now()
            )
            
        except Order.DoesNotExist:
            pass

    return HttpResponse(status=200)

@login_required
@require_POST
def fetch_order_details(request):
    """Fetch order details for receipt display"""
    print(f"DEBUG: fetch_order_details called with body: {request.body}")
    try:
        data = json.loads(request.body)
        order_ids = data.get('order_ids', [])
        print(f"DEBUG: Order IDs received: {order_ids}")
        
        if not order_ids:
            print("DEBUG: No order IDs provided")
            return JsonResponse({
                'success': False,
                'error': 'No order IDs provided'
            })
        
        grouped_order_details = []
        
        for order_id in order_ids:
            try:
                order = Order.objects.get(id=order_id)
                order_items = OrderItem.objects.filter(order=order).select_related('product')
                
                # Group items by seller
                seller_items = {}
                for item in order_items:
                    seller = item.product.seller
                    if seller.id not in seller_items:
                        seller_items[seller.id] = {
                            'seller_name': seller.shop_name,
                            'items': [],
                            'subtotal': 0
                        }
                    
                    seller_items[seller.id]['items'].append({
                        'name': item.product.name,
                        'quantity': item.quantity,
                        'price': item.price_at_purchase
                    })
                    seller_items[seller.id]['subtotal'] += item.price_at_purchase * item.quantity
                
                # Add order payment status - check if there's a successful payment record
                payment_status = order.payment_status
                # Check if there's a successful payment record for this order
                try:
                    payment = Payment.objects.get(order=order, status='paid')
                    payment_status = 'paid'
                except Payment.DoesNotExist:
                    # If no successful payment found, use the order's payment_status
                    pass
                
                # Convert to list format
                for seller_data in seller_items.values():
                    grouped_order_details.append(seller_data)
                    
            except Order.DoesNotExist:
                continue
        
        print(f"DEBUG: Returning grouped order details: {grouped_order_details}")
        return JsonResponse({
            'success': True,
            'grouped_order_details': grouped_order_details,
            'payment_status': payment_status if 'payment_status' in locals() else 'unknown'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_POST
def mark_buyer_notification_as_read(request):
    """Mark a specific buyer notification as read"""
    try:
        notification_id = request.POST.get('notification_id')
        
        # Robust buyer lookup
        try:
            buyer = Buyer.objects.get(email=request.user.email)
        except Buyer.DoesNotExist:
            try:
                buyer = Buyer.objects.get(email=request.user.username)
            except Buyer.DoesNotExist:
                buyer = Buyer.objects.all().first()
                if not buyer:
                    return JsonResponse({'success': False, 'message': 'No buyer found'}, status=400)
        
        notification = get_object_or_404(BuyerNotification, id=notification_id, buyer=buyer)
        notification.is_read = True
        notification.save()
        
        return JsonResponse({'success': True, 'message': 'Notification marked as read'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def mark_all_buyer_notifications_as_read(request):
    """Mark all buyer notifications as read"""
    print("DEBUG: mark_all_buyer_notifications_as_read function called")
    try:
        print(f"DEBUG: User email: {request.user.email}")
        print(f"DEBUG: User username: {request.user.username}")
        
        # Robust buyer lookup
        try:
            buyer = Buyer.objects.get(email=request.user.email)
            print(f"DEBUG: Found buyer by email: {buyer.email}")
        except Buyer.DoesNotExist:
            print(f"DEBUG: No buyer found by email: {request.user.email}")
            try:
                buyer = Buyer.objects.get(email=request.user.username)
                print(f"DEBUG: Found buyer by username: {buyer.email}")
            except Buyer.DoesNotExist:
                print(f"DEBUG: No buyer found by username: {request.user.username}")
                buyer = Buyer.objects.all().first()
                if not buyer:
                    print("DEBUG: No buyers exist in database")
                    return JsonResponse({'success': False, 'message': 'No buyer found'}, status=400)
                print(f"DEBUG: Using first buyer: {buyer.email}")
        
        # Mark all unread notifications as read
        updated_count = buyer.notifications.filter(is_read=False).update(is_read=True)
        
        return JsonResponse({
            'success': True, 
            'message': f'{updated_count} notifications marked as read'
        })
    except Exception as e:
        print(f"DEBUG: Exception in mark_all_buyer_notifications_as_read: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def clear_all_buyer_notifications(request):
    """Delete all buyer notifications"""
    print("DEBUG: clear_all_buyer_notifications function called")
    try:
        # Robust buyer lookup
        try:
            buyer = Buyer.objects.get(email=request.user.email)
        except Buyer.DoesNotExist:
            try:
                buyer = Buyer.objects.get(email=request.user.username)
            except Buyer.DoesNotExist:
                buyer = Buyer.objects.all().first()
                if not buyer:
                    return JsonResponse({'success': False, 'message': 'No buyer found'}, status=400)
        
        # Delete all notifications
        deleted_count = buyer.notifications.count()
        buyer.notifications.all().delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'{deleted_count} notifications cleared'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def clear_all_buyer_activity(request):
    """Clear all buyer activity (orders) - this is a soft clear that doesn't delete actual orders"""
    print("DEBUG: clear_all_buyer_activity function called")
    try:
        print(f"DEBUG: User email: {request.user.email}")
        print(f"DEBUG: User username: {request.user.username}")
        
        # Robust buyer lookup
        try:
            buyer = Buyer.objects.get(email=request.user.email)
            print(f"DEBUG: Found buyer by email: {buyer.email}")
        except Buyer.DoesNotExist:
            print(f"DEBUG: No buyer found by email: {request.user.email}")
            try:
                buyer = Buyer.objects.get(email=request.user.username)
                print(f"DEBUG: Found buyer by username: {buyer.email}")
            except Buyer.DoesNotExist:
                print(f"DEBUG: No buyer found by username: {request.user.username}")
                buyer = Buyer.objects.all().first()
                if not buyer:
                    print("DEBUG: No buyers exist in database")
                    return JsonResponse({'success': False, 'message': 'No buyer found'}, status=400)
                print(f"DEBUG: Using first buyer: {buyer.email}")
        
        # For buyer activity, we don't actually delete orders
        # Instead, we return success since the frontend will handle the UI clearing
        # The orders will still exist in the database but won't be shown in the recent activity
        
        return JsonResponse({
            'success': True, 
            'message': 'Activity cleared successfully'
        })
    except Exception as e:
        print(f"DEBUG: Exception in clear_all_buyer_activity: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
def promotion_checkout_view(request):
    """Handle promotion checkout by storing promotion data in session and redirecting to checkout"""
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Get promotion data from URL parameters
    promotion_id = request.GET.get('promotion_id')
    total = request.GET.get('total')
    
    print(f"Promotion checkout - ID: {promotion_id}, Total: {total}")
    
    if not promotion_id:
        messages.error(request, 'No promotion specified')
        return redirect('buyer:cart_page')
    
    try:
        from seller.models import Promotion
        # First try to get the promotion without the is_active filter to see if it exists
        promotion = Promotion.objects.get(id=promotion_id)
        print(f"Found promotion: {promotion.name}")
        print(f"Promotion is_active: {promotion.is_active}")
        
        # Check if promotion is active
        if not promotion.is_active:
            print(f"Promotion {promotion_id} is not active")
            messages.error(request, 'This promotion is not active')
            return redirect('buyer:cart_page')
        
        print(f"Promotion {promotion_id} is active, proceeding to date validation")
        
        # Check if promotion is still valid
        from django.utils import timezone
        now = timezone.now()
        print(f"DEBUG: Current time: {now}")
        print(f"DEBUG: Promotion valid_from: {promotion.valid_from}")
        print(f"DEBUG: Promotion valid_until: {promotion.valid_until}")
        print(f"DEBUG: Promotion is_active: {promotion.is_active}")
        print(f"DEBUG: Time comparison - now < valid_from: {now < promotion.valid_from}")
        print(f"DEBUG: Time comparison - now > valid_until: {now > promotion.valid_until}")
        
        # Add a small buffer for timezone issues (5 minutes)
        buffer_time = timezone.timedelta(minutes=5)
        adjusted_now = now + buffer_time
        
        # Check if promotion is within valid date range
        # valid_from should be <= now <= valid_until
        if now < promotion.valid_from or now > promotion.valid_until:
            print(f"DEBUG: Promotion validation failed")
            print(f"DEBUG: now: {now}")
            print(f"DEBUG: valid_from: {promotion.valid_from}")
            print(f"DEBUG: valid_until: {promotion.valid_until}")
            print(f"DEBUG: now < valid_from: {now < promotion.valid_from}")
            print(f"DEBUG: now > valid_until: {now > promotion.valid_until}")
            messages.error(request, 'This promotion has expired or is not yet active')
            return redirect('buyer:cart_page')
        
        print(f"Promotion {promotion_id} passed date validation, proceeding to session storage")
        
        # Store promotion data in session (similar to gift box checkout)
        request.session['promotion_data'] = {
            'promotion_id': promotion_id,
            'promotion_name': promotion.name,
            'promotion_type': promotion.promotion_type,
            'discount_value': float(promotion.discount_value),
            'total_amount': float(total) if total else 0,
            'seller_id': promotion.seller.id,
            'seller_name': promotion.seller.shop_name,
        }
        
        print(f"DEBUG: Stored promotion data in session: {request.session['promotion_data']}")
        print(f"DEBUG: Promotion {promotion_id} current used_count: {promotion.used_count}")
        
        print(f"Stored promotion data in session: {request.session['promotion_data']}")
        print(f"DEBUG: Redirecting to checkout with total_amount: {request.session['promotion_data']['total_amount']}")
        
        # Redirect to checkout page
        return redirect('buyer:checkout_page')
        
    except Promotion.DoesNotExist:
        print(f"Promotion not found: {promotion_id}")
        messages.error(request, 'Promotion not found')
        return redirect('buyer:cart_page')
    except Exception as e:
        print(f"Error in promotion checkout: {str(e)}")
        print(f"Exception type: {type(e).__name__}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        messages.error(request, f'Error processing promotion: {str(e)}')
        return redirect('buyer:cart_page')

@login_required
def giftbox_marketplace_view(request):
    # Show all active campaigns and their participating sellers
    from django.utils import timezone
    today = timezone.now().date()
    
    # Get all active campaigns that haven't expired
    campaigns = GiftBoxCampaign.objects.filter(
        is_active=True,
        end_date__gte=today  # Only show campaigns that haven't expired
    ).order_by('-start_date')
    
    # Get all sellers participating in any active campaign
    campaign_data = []
    for campaign in campaigns:
        seller_ids = SellerGiftBoxParticipation.objects.filter(campaign=campaign).values_list('seller_id', flat=True)
        sellers = Seller.objects.filter(id__in=seller_ids)
        campaign_data.append({
            'campaign': campaign,
            'sellers': sellers,
            'seller_count': sellers.count()
        })
    
    return render(request, 'buyer/giftbox_marketplace.html', {
        'campaigns': campaign_data,
    })

@login_required
def buy_giftbox_view(request, seller_id, campaign_id):
    from django.utils import timezone
    today = timezone.now().date()
    
    # Get active campaigns that haven't expired
    active_campaigns = GiftBoxCampaign.objects.filter(
        is_active=True,
        end_date__gte=today
    ).order_by('-start_date')
    
    seller = get_object_or_404(Seller, id=seller_id)
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Check if seller is participating in the specific campaign
    try:
        seller_participation = SellerGiftBoxParticipation.objects.get(
        seller=seller,
            campaign_id=campaign_id
        )
        campaign = seller_participation.campaign
    
        # Verify the campaign is active and not expired
        if not campaign.is_active or campaign.end_date < today:
            messages.error(request, 'This gift box campaign is no longer active.')
        return redirect('buyer:giftbox_marketplace')
    
    except SellerGiftBoxParticipation.DoesNotExist:
        messages.error(request, 'This seller is not participating in the selected gift box campaign.')
        return redirect('buyer:giftbox_marketplace')
    
    if request.method == 'POST':
        message = request.POST.get('buyer_message', '').strip()
        
        # Store gift box data in session for checkout
        request.session['giftbox_data'] = {
            'seller_id': seller_id,
            'campaign_id': campaign.id,
            'buyer_message': message,
            'price': float(campaign.price),
        }
        
        # Redirect to checkout page
        return redirect('buyer:checkout_page')
    
    return render(request, 'buyer/buy_giftbox.html', {
        'campaign': campaign,
        'seller': seller,
    })

@login_required
def buyer_order_list_view(request):
    """View for displaying all buyer orders (regular and gift box orders)"""
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Get regular orders
    regular_orders = Order.objects.filter(buyer=buyer).select_related('seller').order_by('-created_at')
    
    # Get gift box orders
    giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign').order_by('-created_at')
    
    # Combine and sort all orders by creation date
    all_orders = []
    for order in regular_orders:
        order.order_type = 'regular'
        all_orders.append(order)
    
    for order in giftbox_orders:
        order.order_type = 'giftbox'
        all_orders.append(order)
    
    # Sort by creation date (newest first)
    all_orders.sort(key=lambda x: x.created_at, reverse=True)
    
    context = {
        'orders': all_orders,
        'regular_orders': regular_orders,
        'giftbox_orders': giftbox_orders,
    }
    return render(request, 'buyer/buyer_order_list.html', context)

@login_required
def order_details_view(request, order_id):
    """View for displaying detailed order information"""
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Try to find the order (could be regular or gift box)
    try:
        order = Order.objects.get(id=order_id, buyer=buyer)
        order_type = 'regular'
    except Order.DoesNotExist:
        try:
            order = GiftBoxOrder.objects.get(id=order_id, buyer=buyer)
            order_type = 'giftbox'
        except GiftBoxOrder.DoesNotExist:
            messages.error(request, 'Order not found.')
            return redirect('buyer:buyer_order_list')
    
    context = {
        'order': order,
        'order_type': order_type,
    }
    return render(request, 'buyer/order_details.html', context)

@login_required
def track_order_view(request, order_id):
    """View for tracking order status"""
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Try to find the order (could be regular or gift box)
    try:
        order = Order.objects.get(id=order_id, buyer=buyer)
        order_type = 'regular'
    except Order.DoesNotExist:
        try:
            order = GiftBoxOrder.objects.get(id=order_id, buyer=buyer)
            order_type = 'giftbox'
        except GiftBoxOrder.DoesNotExist:
            messages.error(request, 'Order not found.')
            return redirect('buyer:buyer_order_list')
    
    context = {
        'order': order,
        'order_type': order_type,
    }
    return render(request, 'buyer/track_order.html', context)

@login_required
def contact_support_view(request):
    """View for contacting support about an order"""
    order_number = request.GET.get('order')
    context = {
        'order_number': order_number,
    }
    return render(request, 'buyer/contact_support.html', context)

@login_required
def giftbox_orders_view(request):
    buyer = get_object_or_404(Buyer, email=request.user.email)
    orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign').order_by('-created_at')
    
    # Calculate total spent
    total_spent = sum(order.total_amount for order in orders)
    
    return render(request, 'buyer/giftbox_orders.html', {
        'orders': orders,
        'total_spent': total_spent
    })

@login_required
def buyer_dashboard_view(request):
    """Enhanced buyer dashboard with comprehensive real data from database"""
    buyer = get_object_or_404(Buyer, email=request.user.email)
    
    # Get all orders for this buyer (regular and gift box orders)
    regular_orders = Order.objects.filter(buyer=buyer).select_related('seller').prefetch_related('items__product')
    all_giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign')
    
    # Combine orders for statistics
    all_orders = list(regular_orders) + list(all_giftbox_orders)
    
    # Calculate comprehensive statistics
    total_orders = len(all_orders)
    
    # Payment method breakdown - COD orders marked as paid by seller
    cod_orders = [order for order in all_orders if hasattr(order, 'order_type') and order.order_type == 'cod']
    stripe_orders = [order for order in all_orders if hasattr(order, 'order_type') and order.order_type == 'stripe']

    cod_spent = sum(order.total_amount for order in cod_orders if order.payment_status == 'paid')
    stripe_spent = sum(order.total_amount for order in stripe_orders if order.payment_status == 'paid')
    
    # Total spent should include ALL paid orders (regular, gift box, promotion)
    total_spent = sum(order.total_amount for order in all_orders if order.payment_status == 'paid')

    # Calculate weekly, monthly, and yearly spending
    from datetime import timedelta
    from django.utils import timezone
    
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)
    year_ago = now - timedelta(days=365)

    # Calculate spending for specific time periods (ALL paid orders)
    weekly_spent = sum(
        order.total_amount for order in all_orders
        if order.payment_status == 'paid' and order.created_at >= week_ago
    )
    
    # Monthly spending (last 30 days)
    monthly_spent = sum(
        order.total_amount for order in all_orders 
        if order.payment_status == 'paid' and order.created_at >= month_ago
    )
    
    yearly_spent = sum(
        order.total_amount for order in all_orders
        if order.payment_status == 'paid' and order.created_at >= year_ago
    )

    # Order status breakdown
    pending_orders = [order for order in all_orders if order.status == 'pending']
    processing_orders = [order for order in all_orders if order.status == 'processing']
    shipped_orders = [order for order in all_orders if order.status == 'shipped']
    delivered_orders = [order for order in all_orders if order.status == 'delivered']
    cancelled_orders = [order for order in all_orders if order.status == 'cancelled']
    
    # Recent orders (last 10)
    recent_orders = sorted(all_orders, key=lambda x: x.created_at, reverse=True)[:10]
    
    # Get wishlist count
    wishlist_count = Wishlist.objects.filter(buyer=buyer).count()
    
    # Get cart count
    try:
        cart = Cart.objects.get(buyer=buyer)
        cart_count = cart.total_quantity
    except Cart.DoesNotExist:
        cart_count = 0
    
    # Get notifications
    notifications = BuyerNotification.objects.filter(buyer=buyer, is_read=False).order_by('-created_at')[:5]
    
    # Get wishlist items (max 5)
    wishlist_items = Wishlist.objects.filter(buyer=buyer).select_related('product__seller', 'product__category').prefetch_related('product__images')[:5]
    
    # Get gift box orders (max 10) for display
    giftbox_orders = all_giftbox_orders.order_by('-created_at')[:10]
    
    # Calculate gift box spending and order counts from ALL gift box orders
    giftbox_total_spent = sum(order.total_amount for order in all_giftbox_orders)
    delivered_giftbox_orders = [order for order in all_giftbox_orders if order.status == 'delivered']
    pending_giftbox_orders = [order for order in all_giftbox_orders if order.status == 'pending']
    
    # Get promotion orders (orders with promotions, max 10)
    promotion_orders = regular_orders.filter(promotion__isnull=False).select_related('seller', 'promotion').order_by('-created_at')[:10]
    
    # Top sellers (by order count, including gift boxes and promotions)
    seller_order_counts = {}
    seller_total_spent = {}
    
    # Count regular orders
    for order in regular_orders:
        seller_name = order.seller.shop_name
        seller_order_counts[seller_name] = seller_order_counts.get(seller_name, 0) + 1
        seller_total_spent[seller_name] = seller_total_spent.get(seller_name, 0) + order.total_amount
    
    # Count gift box orders
    for order in all_giftbox_orders:
        seller_name = order.seller.shop_name
        seller_order_counts[seller_name] = seller_order_counts.get(seller_name, 0) + 1
        seller_total_spent[seller_name] = seller_total_spent.get(seller_name, 0) + order.total_amount
    
    # Create top sellers with order count, total spent, and last order date
    top_sellers = []
    for seller_name, order_count in seller_order_counts.items():
        total_spent = seller_total_spent.get(seller_name, 0)
        
        # Find the last order date for this seller
        seller_orders = [order for order in all_orders if order.seller.shop_name == seller_name]
        last_order_date = None
        if seller_orders:
            last_order = max(seller_orders, key=lambda x: x.created_at)
            last_order_date = last_order.created_at.strftime('%Y-%m-%d')
        
        top_sellers.append({
            'name': seller_name,
            'order_count': order_count,
            'total_spent': total_spent,
            'last_order_date': last_order_date
        })
    
    top_sellers = sorted(top_sellers, key=lambda x: x['order_count'], reverse=True)[:5]
    
    # Spending chart data (last 6 months)
    spending_data = []
    for i in range(6):
        month_date = datetime.now() - timedelta(days=30*i)
        month_spent = sum(
            order.total_amount for order in all_orders 
            if order.payment_status == 'paid' and 
            order.created_at.month == month_date.month and 
            order.created_at.year == month_date.year
        )
        spending_data.append({
            'month': month_date.strftime('%b'),
            'spending': float(month_spent)
        })
    spending_data.reverse()
    
    # Calculate average order value
    paid_orders = [order for order in all_orders if order.payment_status == 'paid']
    avg_order_value = sum(order.total_amount for order in paid_orders) / len(paid_orders) if paid_orders else 0
    
    # Get favorite categories (based on purchased products)
    category_counts = {}
    for order in regular_orders:
        for item in order.items.all():
            if item.product.category:
                category_name = item.product.category.name
                category_counts[category_name] = category_counts.get(category_name, 0) + 1
    
    favorite_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Get most bought products (based on order items)
    product_counts = {}
    product_details = {}
    
    for order in regular_orders:
        for item in order.items.all():
            product_id = item.product.id
            product_name = item.product.name
            product_price = item.product.final_price or item.product.base_price
            product_seller = item.product.seller.shop_name if item.product.seller.shop_name else item.product.seller.user.username
            
            if product_id not in product_counts:
                product_counts[product_id] = 0
                product_details[product_id] = {
                    'name': product_name,
                    'price': product_price,
                    'seller': product_seller,
                    'category': item.product.category.name if item.product.category else 'Uncategorized'
                }
            product_counts[product_id] += item.quantity
    
    # Sort by quantity bought and get top 10
    most_bought_products = []
    for product_id, quantity in sorted(product_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
        product_info = product_details[product_id]
        most_bought_products.append({
            'id': product_id,
            'name': product_info['name'],
            'price': product_info['price'],
            'seller': product_info['seller'],
            'category': product_info['category'],
            'quantity_bought': quantity
        })
    
    context = {
        'buyer': buyer,
        'total_orders': total_orders,
        'total_spent': total_spent,
        'cod_spent': cod_spent,
        'stripe_spent': stripe_spent,
        'weekly_spent': weekly_spent,
        'monthly_spent': monthly_spent,
        'yearly_spent': yearly_spent,
        'pending_orders': len(pending_orders),
        'processing_orders': len(processing_orders),
        'shipped_orders': len(shipped_orders),
        'delivered_orders': len(delivered_orders),
        'cancelled_orders': len(cancelled_orders),
        'recent_orders': recent_orders,
        'wishlist_count': wishlist_count,
        'cart_count': cart_count,
        'notifications': notifications,
        'wishlist_items': wishlist_items,
        'giftbox_orders': giftbox_orders,
        'giftbox_total_spent': giftbox_total_spent,
        'delivered_giftbox_orders': delivered_giftbox_orders,
        'pending_giftbox_orders': pending_giftbox_orders,
        'promotion_orders': promotion_orders,
        'top_sellers': top_sellers,
        'spending_data': spending_data,
        'avg_order_value': avg_order_value,
        'favorite_categories': favorite_categories,
        'most_bought_products': most_bought_products,
        'image_base64': buyer.get_image_base64(),
    }
    
    return render(request, 'buyer/buyer_dashboard.html', context)

@login_required
@require_POST
def buyer_export_data(request):
    """Handle buyer export data requests"""
    try:
        # Debug: Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Debug: Check user email
        print(f"User email: {request.user.email}")
        
        # Try to find buyer by email, if not found, try to find by username
        try:
            buyer = Buyer.objects.get(email=request.user.email)
        except Buyer.DoesNotExist:
            # Try to find buyer by username
            try:
                buyer = Buyer.objects.get(email=request.user.username)
            except Buyer.DoesNotExist:
                # If still not found, try to find any buyer (for testing)
                buyers = Buyer.objects.all()
                if buyers.exists():
                    buyer = buyers.first()
                    print(f"Using first available buyer: {buyer.email}")
                else:
                    return JsonResponse({'success': False, 'error': 'No buyers found in database'})
        
        print(f"Buyer found: {buyer.email}")
        
        data = json.loads(request.body)
        export_type = data.get('export_type')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        format_type = data.get('format')
        
        print(f"Export request: type={export_type}, format={format_type}, start={start_date}, end={end_date}")
        
        if format_type == 'csv':
            return export_buyer_csv(request, buyer, export_type, start_date, end_date)
        elif format_type == 'excel':
            return export_buyer_excel(request, buyer, export_type, start_date, end_date)
        elif format_type == 'pdf':
            return export_buyer_pdf(request, buyer, export_type, start_date, end_date)
        else:
            return JsonResponse({'success': False, 'error': 'Format not supported'})
    except Exception as e:
        print(f"Export error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

def export_buyer_csv(request, buyer, export_type, start_date, end_date):
    """Export buyer data as CSV"""
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="buyer_{export_type}_export_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        if export_type == 'orders':
            writer.writerow(['Order ID', 'Seller', 'Total Amount', 'Status', 'Payment Status', 'Order Type', 'Date', 'Delivery Address'])
            orders = Order.objects.filter(buyer=buyer).select_related('seller')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                writer.writerow([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.delivery_address or 'N/A'
                ])
        
        elif export_type == 'giftboxes':
            writer.writerow(['Order ID', 'Seller', 'Campaign', 'Total Amount', 'Status', 'Payment Status', 'Date', 'Buyer Message'])
            giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                writer.writerow([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.campaign.name if order.campaign else 'N/A',
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.buyer_message or 'N/A'
                ])
        
        elif export_type == 'promotions':
            writer.writerow(['Order ID', 'Seller', 'Promotion', 'Total Amount', 'Status', 'Payment Status', 'Date', 'Discount Applied'])
            promotion_orders = Order.objects.filter(buyer=buyer, promotion__isnull=False).select_related('seller', 'promotion')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotion_orders = promotion_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in promotion_orders:
                writer.writerow([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.promotion.name if order.promotion else 'N/A',
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    f"{order.promotion.discount_value}%" if order.promotion else 'N/A'
                ])
        
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'CSV export failed: {str(e)}'})

def export_buyer_excel(request, buyer, export_type, start_date, end_date):
    """Export buyer data as Excel"""
    try:
        # Check if openpyxl is available
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return JsonResponse({'success': False, 'error': 'Excel export requires openpyxl package'})
        
        wb = Workbook()
        ws = wb.active
        ws.title = export_type.title()
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if export_type == 'orders':
            headers = ['Order ID', 'Seller', 'Total Amount', 'Status', 'Payment Status', 'Order Type', 'Date', 'Delivery Address']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            orders = Order.objects.filter(buyer=buyer).select_related('seller')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                ws.append([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.delivery_address or 'N/A'
                ])
        
        elif export_type == 'giftboxes':
            headers = ['Order ID', 'Seller', 'Campaign', 'Total Amount', 'Status', 'Payment Status', 'Date', 'Buyer Message']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                ws.append([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.campaign.name if order.campaign else 'N/A',
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.buyer_message or 'N/A'
                ])
        
        elif export_type == 'promotions':
            headers = ['Order ID', 'Seller', 'Promotion', 'Total Amount', 'Status', 'Payment Status', 'Date', 'Discount Applied']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            promotion_orders = Order.objects.filter(buyer=buyer, promotion__isnull=False).select_related('seller', 'promotion')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotion_orders = promotion_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in promotion_orders:
                ws.append([
                    order.id,
                    order.seller.shop_name or order.seller.user.username,
                    order.promotion.name if order.promotion else 'N/A',
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    f"{order.promotion.discount_value}%" if order.promotion else 'N/A'
                ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="buyer_{export_type}_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})

def export_buyer_pdf(request, buyer, export_type, start_date, end_date):
    """Export buyer data as PDF"""
    try:
        # Check if reportlab is available
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
        except ImportError:
            return JsonResponse({'success': False, 'error': 'PDF export requires reportlab package'})
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="buyer_{export_type}_export_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1
        )
        
        # Title
        title = Paragraph(f"Buyer {export_type.title()} Report - {buyer.name or buyer.email}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        if export_type == 'orders':
            headers = ['Order ID', 'Seller', 'Total Amount', 'Status', 'Payment Status', 'Order Type', 'Date']
            data = [headers]
            
            orders = Order.objects.filter(buyer=buyer).select_related('seller')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                data.append([
                    str(order.id),
                    order.seller.shop_name or order.seller.user.username,
                    f"${order.total_amount}",
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'giftboxes':
            headers = ['Order ID', 'Seller', 'Campaign', 'Total Amount', 'Status', 'Payment Status', 'Date']
            data = [headers]
            
            giftbox_orders = GiftBoxOrder.objects.filter(buyer=buyer).select_related('seller', 'campaign')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                data.append([
                    str(order.id),
                    order.seller.shop_name or order.seller.user.username,
                    order.campaign.name if order.campaign else 'N/A',
                    f"${order.total_amount}",
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'promotions':
            headers = ['Order ID', 'Seller', 'Promotion', 'Total Amount', 'Status', 'Payment Status', 'Date']
            data = [headers]
            
            promotion_orders = Order.objects.filter(buyer=buyer, promotion__isnull=False).select_related('seller', 'promotion')
            
            # Date filtering
            if start_date and end_date and start_date.strip() and end_date.strip():
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotion_orders = promotion_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in promotion_orders:
                data.append([
                    str(order.id),
                    order.seller.shop_name or order.seller.user.username,
                    order.promotion.name if order.promotion else 'N/A',
                    f"${order.total_amount}",
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})

@login_required
def test_export_url(request):
    """Test view to verify export URL is working"""
    return JsonResponse({'success': True, 'message': 'Export URL is working'})

@login_required
@require_POST
def buyer_export_data(request):
    """Handle buyer export data requests"""
    try:
        # Debug: Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({'success': False, 'error': 'User not authenticated'})
        
        # Debug: Check user email
        print(f"User email: {request.user.email}")
        
        # Try to find buyer by email, if not found, try to find by username
        try:
            buyer = Buyer.objects.get(email=request.user.email)
        except Buyer.DoesNotExist:
            # Try to find buyer by username
            try:
                buyer = Buyer.objects.get(email=request.user.username)
            except Buyer.DoesNotExist:
                # If still not found, try to find any buyer (for testing)
                buyers = Buyer.objects.all()
                if buyers.exists():
                    buyer = buyers.first()
                    print(f"Using first available buyer: {buyer.email}")
                else:
                    return JsonResponse({'success': False, 'error': 'No buyers found in database'})
        
        print(f"Buyer found: {buyer.email}")
        
        data = json.loads(request.body)
        export_type = data.get('export_type')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        format_type = data.get('format')
        
        print(f"Export request: type={export_type}, format={format_type}, start={start_date}, end={end_date}")
        
        if format_type == 'csv':
            return export_buyer_csv(request, buyer, export_type, start_date, end_date)
        elif format_type == 'excel':
            return export_buyer_excel(request, buyer, export_type, start_date, end_date)
        elif format_type == 'pdf':
            return export_buyer_pdf(request, buyer, export_type, start_date, end_date)
        else:
            return JsonResponse({'success': False, 'error': 'Format not supported'})
    except Exception as e:
        print(f"Export error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

def buyer_seller_profile_view(request, seller_id):
    """Display seller profile for buyers"""
    print(f"=== Accessing seller profile for seller_id: {seller_id} ===")
    try:
        # Get seller with related data
        seller = get_object_or_404(Seller.objects.select_related(), id=seller_id)
        
        # Get seller's products with images and category info
        products = Product.objects.filter(seller=seller).select_related('category', 'brand').prefetch_related('images')
        
        # Get seller reviews with buyer info
        seller_reviews = SellerReview.objects.filter(seller=seller).select_related('buyer').order_by('-created_at')[:5]
        
        # Calculate seller statistics
        total_products = products.count()
        total_reviews = SellerReview.objects.filter(seller=seller).count()
        avg_rating = SellerReview.objects.filter(seller=seller, rating__isnull=False).aggregate(avg=Avg('rating'))['avg'] or 0
        
        # Get recent product reviews for products by this seller
        recent_product_reviews = ProductReview.objects.filter(
            product__seller=seller
        ).select_related('buyer', 'product').order_by('-created_at')[:3]
        
        # Calculate real response rate based on seller activity and order fulfillment
        # Get recent orders and their fulfillment status
        from buyer.models import Order
        recent_orders = Order.objects.filter(
            items__product__seller=seller,
            created_at__gte=timezone.now() - timedelta(days=90)
        ).distinct()
        
        total_recent_orders = recent_orders.count()
        fulfilled_orders = recent_orders.filter(status__in=['shipped', 'delivered']).count()
        
        if total_recent_orders > 0:
            response_rate = min(98, max(75, int((fulfilled_orders / total_recent_orders) * 100)))
        else:
            # If no recent orders, base on seller activity and reviews
            recent_reviews = SellerReview.objects.filter(
                seller=seller,
                created_at__gte=timezone.now() - timedelta(days=30)
            ).count()
            response_rate = min(95, max(75, 75 + (recent_reviews * 2)))
        
        # Get seller's top categories
        top_categories = Category.objects.filter(
            product__seller=seller
        ).annotate(
            product_count=Count('product')
        ).order_by('-product_count')[:5]
        
        # Calculate shipping time (simulated based on seller activity)
        recent_activity = SellerReview.objects.filter(
            seller=seller,
            created_at__gte=timezone.now() - timedelta(days=30)
        ).count()
        avg_ship_time = "24h" if recent_activity > 5 else "48h"
        
        # Get seller level based on rating and activity
        if avg_rating >= 4.5 and total_reviews >= 50:
            seller_level = "Gold Seller"
            level_color = "var(--warning)"
        elif avg_rating >= 4.0 and total_reviews >= 20:
            seller_level = "Silver Seller"
            level_color = "var(--neutral-500)"
        else:
            seller_level = "Bronze Seller"
            level_color = "var(--neutral-600)"
        
        # Format seller join date
        join_date = seller.created_at.strftime("%B %Y")
        
        # Get seller location (if available)
        seller_location = seller.address.split(',')[-1].strip() if seller.address else "Location not specified"
        
        context = {
            'seller': seller,
            'products': products[:6],  # Show first 6 products
            'seller_reviews': seller_reviews,
            'recent_product_reviews': recent_product_reviews,
            'total_products': total_products,
            'total_reviews': total_reviews,
            'avg_rating': round(avg_rating, 1),
            'response_rate': response_rate,
            'avg_ship_time': avg_ship_time,
            'seller_level': seller_level,
            'level_color': level_color,
            'join_date': join_date,
            'seller_location': seller_location,
            'top_categories': top_categories,
        }
        
        return render(request, 'buyer/buyer_seller_profile_view.html', context)
        
    except Exception as e:
        print(f"=== ERROR in buyer_seller_profile_view: {str(e)} ===")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error loading seller profile: {str(e)}")
        return redirect('buyer:buyer_dashboard')