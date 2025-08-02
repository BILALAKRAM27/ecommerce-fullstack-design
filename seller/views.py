import base64
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import login, authenticate, logout
from django.core.paginator import Paginator
from .models import Seller, Product, ProductImage, Brand, Category, CategoryAttribute, AttributeOption, ProductAttributeValue, Notification, Activity
from .forms import UserRegisterForm, SellerUpdateForm, SellerLoginForm, ProductForm, ProductImageForm, DynamicProductForm
from django.utils import timezone
from buyer.models import Buyer, Order, OrderStatus, Payment, PaymentStatus
from django.http import HttpResponse, JsonResponse, HttpResponseBadRequest
from django.views.decorators.csrf import csrf_exempt
import json
import os
import stripe
from django.conf import settings
from django.shortcuts import redirect
from .models import Seller
from django.views.decorators.http import require_POST, require_GET
from django.db.models import Sum, Count, Q, Min, Max, F
from datetime import datetime, timedelta
import csv
from io import StringIO
from buyer.models import OrderItem
import logging
from .models import Promotion
import csv
from django.utils import timezone
from django.http import HttpResponse
from seller.models import Product
from datetime import datetime
from .models import GiftBoxCampaign, SellerGiftBoxParticipation
from django.urls import reverse
from buyer.models import GiftBoxOrder
from django.forms import ModelMultipleChoiceField
from django.http import JsonResponse
from .models import ProductReview, ProductReviewLike, SellerReview, SellerReviewLike


stripe.api_key = os.getenv("STRIPE_SECRET_KEY") or settings.STRIPE_SECRET_KEY


def index_view(request):
    user = request.user if request.user.is_authenticated else None
    seller = None
    buyer = None
    image_base64 = None
    user_type = None
    if user:
        # Try to get Seller by user
        try:
            seller = Seller.objects.get(user=user)
            image_base64 = seller.get_image_base64()
            user_type = 'seller'
        except Seller.DoesNotExist:
            # Try to get Buyer by email
            try:
                buyer = Buyer.objects.get(email=user.email)
                image_base64 = buyer.get_image_base64()
                user_type = 'buyer'
            except Buyer.DoesNotExist:
                pass
    # Get categories for sections with improved logic
    electronics = Category.objects.filter(name__icontains='electronic')
    electronics_subs = Category.objects.filter(parent__in=electronics)
    clothing = Category.objects.filter(name__icontains='clothing')
    clothing_subs = Category.objects.filter(parent__in=clothing)
    furniture = Category.objects.filter(name__icontains='furniture')
    furniture_subs = Category.objects.filter(parent__in=furniture)
    sports = Category.objects.filter(name__icontains='sport')
    sports_subs = Category.objects.filter(parent__in=sports)
    electronics_cats = list(electronics) + list(electronics_subs)
    clothing_cats = list(clothing) + list(clothing_subs)
    furniture_cats = list(furniture) + list(furniture_subs)
    sports_cats = list(sports) + list(sports_subs)
    
    # Fetch top 6 trending products for each section based on order_count and rating
    electronics_products = Product.objects.filter(
        category__in=electronics_cats
    ).annotate(
        popularity_score=F('order_count') + (F('rating_avg') * 10)
    ).order_by('-popularity_score', '-order_count', '-rating_avg')[:6]
    
    clothing_products = Product.objects.filter(
        category__in=clothing_cats
    ).annotate(
        popularity_score=F('order_count') + (F('rating_avg') * 10)
    ).order_by('-popularity_score', '-order_count', '-rating_avg')[:6]
    
    furniture_products = Product.objects.filter(
        category__in=furniture_cats
    ).annotate(
        popularity_score=F('order_count') + (F('rating_avg') * 10)
    ).order_by('-popularity_score', '-order_count', '-rating_avg')[:6]
    
    sports_products = Product.objects.filter(
        category__in=sports_cats
    ).annotate(
        popularity_score=F('order_count') + (F('rating_avg') * 10)
    ).order_by('-popularity_score', '-order_count', '-rating_avg')[:6]
    
    # Get active promotions for homepage
    active_promotions = get_active_promotions_for_homepage()
    
    # Get the soonest expiring promotion for the timer
    soonest_expiring_promotion = None
    if active_promotions:
        soonest_expiring_promotion = active_promotions.first()
    
    # Get all parent categories for sidebar
    parent_categories = Category.objects.filter(parent__isnull=True).order_by('name')
    
    # Get recommended items based on highest order count and rating
    recommended_items = Product.objects.annotate(
        recommendation_score=F('order_count') + (F('rating_avg') * 10)
    ).filter(
        order_count__gt=0  # Only products that have been ordered
    ).order_by('-recommendation_score', '-order_count', '-rating_avg')[:10]
    
    context = {
        'user': user,
        'seller': seller,
        'buyer': buyer,
        'image_base64': image_base64,
        'user_type': user_type,
        'parent_categories': parent_categories,
        'electronics_categories': electronics_cats,
        'clothing_categories': clothing_cats,
        'furniture_categories': furniture_cats,
        'electronics_products': electronics_products,
        'clothing_products': clothing_products,
        'furniture_products': furniture_products,
        'sports_products': sports_products,
        'recommended_items': recommended_items,
        'active_promotions': active_promotions,
        'soonest_expiring_promotion': soonest_expiring_promotion,
    }
    return render(request, "index.html", context)

# CREATE User (Buyer or Seller)
def create_seller_view(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST, request.FILES)
        if form.is_valid():
            # Create the User account
            user = form.save()
            user_type = form.cleaned_data['user_type']
            
            if user_type == 'seller':
                # Create the Seller profile
                seller = Seller.objects.create(
                    user=user,
                    name=user.username,  # Use username as name
                    email=user.email,
                    shop_name=form.cleaned_data.get('shop_name', ''),
                    shop_description='',
                    address='',  # Empty address initially
                    created_at=timezone.now()
                )
                
                messages.success(request, 'Seller account created successfully!')
            else:
                # Create the Buyer profile
                buyer = Buyer.objects.create(
                    name=user.username,  # Use username as name
                    email=user.email,
                    phone=form.cleaned_data.get('phone', ''),
                    address='',  # Empty address
                    created_at=timezone.now()
                )
                
                messages.success(request, 'Buyer account created successfully!')
            
            return redirect('sellers:login')
    else:
        form = UserRegisterForm()
    
    return render(request, 'seller/register.html', {'form': form, 'title': 'Register'})


# LOGIN View
def login_view(request):
    if request.method == 'POST':
        form = SellerLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            user_type = form.cleaned_data['user_type']
            
            # Log the user in
            login(request, user)
            
            messages.success(request, f'Welcome back, {user.username}!')
            
            if user_type == 'seller':
                return redirect('sellers:seller_profile')
            else:
                # Check if buyer exists
                try:
                    buyer = Buyer.objects.get(email=user.email)
                    return redirect('buyer:buyer_profile')  # You'll need to create this view
                except Buyer.DoesNotExist:
                    messages.error(request, 'This email is not registered as a buyer.')
                    logout(request)
                    return redirect('sellers:login')
    else:
        form = SellerLoginForm()
    
    return render(request, 'seller/login.html', {'form': form, 'title': 'Login'})


# LOGOUT View
@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('sellers:login')


# READ Seller Profile
@login_required
def seller_profile_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        # Handle form submission
        try:
            # Update basic fields
            seller.name = request.POST.get('name', seller.name)
            seller.shop_name = request.POST.get('shop_name', seller.shop_name)
            seller.shop_description = request.POST.get('shop_description', seller.shop_description)
            seller.address = request.POST.get('address', seller.address)
            
            # Handle image upload
            if 'image' in request.FILES:
                image_file = request.FILES['image']
                # Read the image file and store as binary data
                seller.image = image_file.read()
            
            seller.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('sellers:seller_profile')
            
        except Exception as e:
            messages.error(request, f'Error updating profile: {str(e)}')
    
    context = {
        'seller': seller,
        'image_base64': seller.get_image_base64()
    }
    return render(request, 'seller/seller_profile.html', context)


# READ Buyer Profile



# UPDATE Seller
@login_required
def update_seller_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    if request.method == 'POST':
        form = SellerUpdateForm(request.POST, request.FILES, instance=seller)
        if form.is_valid():
            seller = form.save()  # Let the form handle image upload
            seller.updated_at = timezone.now()
            seller.save()
            messages.success(request, 'Seller profile updated successfully.')
            return redirect('sellers:seller_profile')
    else:
        form = SellerUpdateForm(instance=seller)
    return render(request, 'seller/seller_update.html', {'form': form})


# DELETE Seller
@login_required
def delete_seller_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    if request.method == 'POST':
        # Store the user reference before deleting seller
        user = seller.user
        
        # Delete the seller profile
        seller.delete()
        
        # Delete the associated user account
        if user:
            user.delete()
        
        messages.success(request, 'Your seller account has been deleted.')
        return redirect('sellers:login')
    return render(request, 'seller/seller_delete.html')


# ========== PRODUCT CRUD VIEWS ==========

# CREATE Product
@login_required
def create_product_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        form = DynamicProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.seller = seller
            
            # Auto-calculate final price based on discount
            base_price = product.base_price or 0
            discount_percentage = product.discount_percentage or 0
            
            if discount_percentage > 0:
                # Calculate final price with discount
                discount_amount = base_price * (discount_percentage / 100)
                product.final_price = max(0, base_price - discount_amount)
            else:
                # If no discount, final price equals base price
                product.final_price = base_price
            
            product.save()
            
            # Handle dynamic attributes (this will be called by the form's save method)
            form.save_dynamic_attributes(product)
            
            # Handle new images
            thumbnail_index = int(request.POST.get('thumbnail', 0))
            image_files = request.FILES.getlist('image_file')
            for idx, image_file in enumerate(image_files):
                ProductImage.objects.create(
                    product=product,
                    image=image_file.read(),
                    is_thumbnail=(idx == thumbnail_index)
                )
            # Handle existing images (if editing)
            existing_thumbnail_id = request.POST.get('existing_thumbnail')
            if existing_thumbnail_id:
                for img in product.images.all():
                    img.is_thumbnail = (str(img.id) == existing_thumbnail_id)
                    img.save()
            messages.success(request, 'Product created successfully!')
            return redirect('seller:products_list')
    else:
        form = DynamicProductForm()
    
    # Get parent categories (categories with no parent)
    parent_categories = Category.objects.filter(parent__isnull=True)
    
    context = {
        'form': form,
        'title': 'Add New Product',
        'seller': seller,
        'parent_categories': parent_categories
    }
    return render(request, 'seller/product_form.html', context)


# READ Product List


# READ Product Detail
def product_page_view(request, product_id):
    user = request.user if request.user.is_authenticated else None
    seller = None
    user_type = None
    in_wishlist = False
    if user:
        try:
            seller = Seller.objects.get(user=user)
            user_type = 'seller'
        except Seller.DoesNotExist:
            from buyer.models import Buyer, Wishlist
            try:
                buyer = Buyer.objects.get(email=user.email)
                user_type = 'buyer'
                # Check if product is in wishlist
                in_wishlist = Wishlist.objects.filter(buyer=buyer, product_id=product_id).exists()
            except Buyer.DoesNotExist:
                pass
    product = get_object_or_404(Product, id=product_id)
    context = {
        'product': product,
        'seller': seller,
        'user_type': user_type,
        'in_wishlist': in_wishlist,
        # add 'user' if you use it in the template
    }
    return render(request, 'seller/product_page.html', context)


# UPDATE Product (for seller dashboard - AJAX)
@login_required
def update_product_view(request, product_id):
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('x-requested-with') == 'XMLHttpRequest'
        
        try:
            # Use form validation for all requests (like the working repo code)
            form = DynamicProductForm(request.POST, request.FILES, instance=product)
            if form.is_valid():
                product = form.save(commit=False)
                
                # Auto-calculate final price based on discount
                base_price = product.base_price or 0
                discount_percentage = product.discount_percentage or 0
                
                if discount_percentage > 0:
                    # Calculate final price with discount
                    discount_amount = base_price * (discount_percentage / 100)
                    product.final_price = max(0, base_price - discount_amount)
                else:
                    # If no discount, final price equals base price
                    product.final_price = base_price
                
                product.save()
            else:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'Form validation failed'})
                else:
                    messages.error(request, 'Form validation failed!')
                    return redirect('sellers:product_page', product_id=product.id)
            
            # --- Image upload logic (exact working code from repo) ---
            image_files = request.FILES.getlist('image_file')
            image_mode = request.POST.get('image_mode', 'add')
            
            if image_files:
                if image_mode == 'replace':
                    # Delete all existing images for this product
                    product.images.all().delete()
                # Handle new images
                thumbnail_index = int(request.POST.get('thumbnail', 0))
                for idx, image_file in enumerate(image_files):
                    image_data = image_file.read()
                    ProductImage.objects.create(
                        product=product,
                        image=image_data,
                        is_thumbnail=(idx == thumbnail_index)
                    )
            # Handle existing images (if editing)
            existing_thumbnail_id = request.POST.get('existing_thumbnail')
            if existing_thumbnail_id:
                for img in product.images.all():
                    img.is_thumbnail = (str(img.id) == existing_thumbnail_id)
                    img.save()
            if existing_thumbnail_id:
                for img in product.images.all():
                    img.is_thumbnail = (str(img.id) == existing_thumbnail_id)
                    img.save()
            
            # Return JSON for AJAX requests, redirect for regular requests
            if is_ajax:
                return JsonResponse({
                    'success': True, 
                    'message': 'Product updated successfully',
                    'product': {
                        'id': product.id,
                        'name': product.name,
                        'description': product.description,
                        'base_price': product.base_price,
                        'final_price': product.final_price,
                        'stock': product.stock,
                        'condition': product.condition,
                        'discount_percentage': product.discount_percentage
                    }
                })
            else:
                messages.success(request, 'Product updated successfully!')
                return redirect('sellers:product_page', product_id=product.id)
                
        except Exception as e:
            if is_ajax:
                return JsonResponse({'success': False, 'error': str(e)})
            else:
                messages.error(request, f'Error updating product: {str(e)}')
                return redirect('sellers:product_page', product_id=product.id)
    
    # For GET requests, render the form (non-AJAX)
    form = DynamicProductForm(instance=product)
    
    # Get parent categories (categories with no parent)
    parent_categories = Category.objects.filter(parent__isnull=True)
    
    # Get current category hierarchy for pre-selection
    current_category = product.category
    parent_category = None
    child_category = None
    
    if current_category.parent:
        parent_category = current_category.parent
        child_category = current_category
    else:
        parent_category = current_category
    
    # Get existing attribute values for pre-population
    existing_attributes = {}
    for attr_value in product.attribute_values.all():
        existing_attributes[f'attribute_{attr_value.attribute.id}'] = attr_value.value
    
    context = {
        'form': form,
        'product': product,
        'title': 'Update Product',
        'seller': seller,
        'parent_categories': parent_categories,
        'current_parent_category': parent_category,
        'current_child_category': child_category,
        'existing_attributes': existing_attributes
    }
    return render(request, 'seller/product_form.html', context)





# AJAX Update Product Form (for seller dashboard)
@login_required
def update_product_form_view(request, product_id):
    """AJAX view to load update form in modal"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    form = DynamicProductForm(instance=product)
    
    # Get parent categories (categories with no parent)
    parent_categories = Category.objects.filter(parent__isnull=True)
    
    # Get current category hierarchy for pre-selection
    current_category = product.category
    parent_category = None
    child_category = None
    
    if current_category.parent:
        parent_category = current_category.parent
        child_category = current_category
    else:
        parent_category = current_category
    
    # Get existing attribute values for pre-population
    existing_attributes = {}
    for attr_value in product.attribute_values.all():
        existing_attributes[f'attribute_{attr_value.attribute.id}'] = attr_value.value
    
    context = {
        'form': form,
        'product': product,
        'seller': seller,
        'parent_categories': parent_categories,
        'current_parent_category': parent_category,
        'current_child_category': child_category,
        'existing_attributes': existing_attributes,
        'is_update': True,
        'form_title': 'Update Product',
        'submit_text': 'Update Product'
    }
    return render(request, 'seller/product_form_modal.html', context)





# DELETE Product
@login_required
def delete_product_view(request, product_id):
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('seller:products_list')
    
    context = {
        'product': product,
        'seller': seller
    }
    return render(request, 'seller/product_delete.html', context)


# ========== AJAX VIEWS FOR DYNAMIC FORM ==========

@csrf_exempt
def get_category_children(request):
    """Get child categories for a selected parent category"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            parent_id = data.get('parent_id')
            
            if parent_id:
                children = Category.objects.filter(parent_id=parent_id)
                children_data = [{'id': child.id, 'name': child.name} for child in children]
                return JsonResponse({'success': True, 'children': children_data})
            else:
                return JsonResponse({'success': False, 'error': 'Parent ID is required'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def get_category_attributes(request):
    """Get attributes for a selected category with existing options"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            category_id = data.get('category_id')
            
            if category_id:
                attributes = CategoryAttribute.objects.filter(category_id=category_id)
                attributes_data = []
                
                for attr in attributes:
                    attr_data = {
                        'id': attr.id,
                        'name': attr.name,
                        'input_type': attr.input_type,
                        'is_required': attr.is_required,
                        'unit': attr.unit,
                        'options': []
                    }
                    
                    # Get existing options for all attribute types (not just dropdown)
                    options = AttributeOption.objects.filter(attribute=attr)
                    attr_data['options'] = [{'id': opt.id, 'value': opt.value} for opt in options]
                    
                    attributes_data.append(attr_data)
                
                return JsonResponse({'success': True, 'attributes': attributes_data})
            else:
                return JsonResponse({'success': False, 'error': 'Category ID is required'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def save_attribute_value(request):
    """Save attribute value and create option if new"""
    if request.method == 'POST':
        data = json.loads(request.body)
        attribute_id = data.get('attribute_id')
        value = data.get('value')
        
        if attribute_id and value:
            try:
                attribute = CategoryAttribute.objects.get(id=attribute_id)
                
                # Check if this value already exists as an option
                existing_option = AttributeOption.objects.filter(
                    attribute=attribute, 
                    value=value
                ).first()
                
                if not existing_option:
                    # Create new option
                    AttributeOption.objects.create(
                        attribute=attribute,
                        value=value
                    )
                
                return JsonResponse({'success': True})
            except CategoryAttribute.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Attribute not found'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@login_required
def stripe_onboard(request):
    seller = get_object_or_404(Seller, user=request.user)
    stripe.api_key = settings.STRIPE_SECRET_KEY

    if not seller.stripe_account_id:
        # Create a new Stripe account for the seller
        account = stripe.Account.create(
            type="express",
            country="US",
            email=seller.email,
            capabilities={"transfers": {"requested": True}},
        )
        seller.stripe_account_id = account.id
        seller.save()
    # Create an onboarding link
    account_link = stripe.AccountLink.create(
        account=seller.stripe_account_id,
        refresh_url=request.build_absolute_uri('/seller/stripe/refresh/'),
        return_url=request.build_absolute_uri('/seller/dashboard/'),
        type="account_onboarding",
    )
    return redirect(account_link.url)

@login_required
def seller_dashboard(request):
    """Seller dashboard overview with order history and revenue"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Get all orders for this seller
    orders = Order.objects.filter(seller=seller).order_by('-created_at')
    
    # Calculate revenue statistics
    total_orders = orders.count()
    total_revenue = sum(order.total_amount for order in orders if order.payment_status == 'paid')
    pending_orders = orders.filter(status='pending').count()
    completed_orders = orders.filter(status='delivered').count()
    
    # Recent orders (last 10)
    recent_orders = orders[:10]
    
    context = {
        'seller': seller,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'recent_orders': recent_orders,
    }
    return render(request, 'seller/seller_dashboard.html', context)

@login_required
def seller_orders(request):
    """Detailed order list for seller"""
    seller = get_object_or_404(Seller, user=request.user)
    orders = Order.objects.filter(seller=seller).order_by('-created_at')
    
    context = {
        'seller': seller,
        'orders': orders,
    }
    return render(request, 'seller/orders.html', context)

@login_required
def seller_revenue(request):
    """Revenue and payout tracking for seller"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Get all paid orders
    paid_orders = Order.objects.filter(seller=seller, payment_status='paid')
    
    # Calculate revenue by month
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth
    
    monthly_revenue = paid_orders.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Sum('total_amount')
    ).order_by('month')
    
    # Calculate commission (10% of total revenue)
    total_revenue = sum(order.total_amount for order in paid_orders)
    total_commission = total_revenue * 0.10
    net_payout = total_revenue - total_commission
    
    context = {
        'seller': seller,
        'total_revenue': total_revenue,
        'total_commission': total_commission,
        'net_payout': net_payout,
        'monthly_revenue': monthly_revenue,
        'paid_orders': paid_orders,
    }
    return render(request, 'seller/revenue.html', context)

@login_required
def seller_stripe_status(request):
    """Check seller's Stripe onboarding status"""
    seller = get_object_or_404(Seller, user=request.user)
    
    context = {
        'seller': seller,
        'has_stripe': bool(seller.stripe_account_id),
    }
    return render(request, 'seller/stripe_status.html', context)

@login_required
def seller_dashboard_overview(request):
    """Enhanced seller dashboard with comprehensive real data from database"""
    # Allow superuser to select seller by ?seller_id= for debugging
    if request.user.is_superuser and request.GET.get('seller_id'):
        try:
            seller = Seller.objects.get(id=request.GET['seller_id'])
        except Seller.DoesNotExist:
            return HttpResponseBadRequest('Seller not found')
    else:
        seller = get_object_or_404(Seller, user=request.user)

    # Debug logging
    logger = logging.getLogger(__name__)
    logger.info(f"Dashboard for seller ID: {seller.id}, shop: {seller.shop_name}")

    # Get all orders for this seller with detailed information
    all_orders = Order.objects.filter(seller=seller).select_related('buyer').prefetch_related('items__product')
    logger.info(f"Order count: {all_orders.count()}")

    # Get all products for this seller
    products = Product.objects.filter(seller=seller)
    logger.info(f"Product count: {products.count()}")

    # Get recent orders (last 10) with buyer information
    recent_orders = all_orders.order_by('-created_at')[:10]

    # Annotate products for dashboard
    products = products.prefetch_related('images').annotate(
        dashboard_order_count=Count('orderitem__order', distinct=True),
        dashboard_total_sold=Sum('orderitem__quantity'),
        dashboard_total_revenue=Sum('orderitem__price_at_purchase')
    ).order_by('-created_at')

    # Calculate comprehensive statistics
    total_orders = all_orders.count()
    total_products = products.count()

    # Payment method breakdown - use Order model's order_type field
    cod_revenue = all_orders.filter(
        order_type='cod',
        payment_status='paid',
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    stripe_revenue = all_orders.filter(
        order_type='stripe',
        payment_status='paid'
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Total earnings = COD revenue (paid + delivered) + Stripe revenue (paid)
    total_earnings = cod_revenue + stripe_revenue

    # Monthly sales (current month) - COD (paid + delivered) + Stripe (paid)
    current_month = datetime.now().month
    current_year = datetime.now().year
    
    # Monthly COD revenue (paid + delivered orders in current month)
    monthly_cod_revenue = all_orders.filter(
        order_type='cod',
        payment_status='paid',
        status='delivered',
        created_at__month=current_month,
        created_at__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Monthly Stripe revenue (paid orders in current month)
    monthly_stripe_revenue = all_orders.filter(
        order_type='stripe',
        payment_status='paid',
        created_at__month=current_month,
        created_at__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_sales = monthly_cod_revenue + monthly_stripe_revenue

    # Revenue breakdown by time periods
    weekly_revenue = all_orders.filter(
        payment_status='paid',
        created_at__gte=datetime.now() - timedelta(days=7)
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    monthly_revenue = all_orders.filter(
        payment_status='paid',
        created_at__gte=datetime.now() - timedelta(days=30)
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    yearly_revenue = all_orders.filter(
        payment_status='paid',
        created_at__year=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # Order status breakdown
    pending_orders = all_orders.filter(status='pending').count()
    processing_orders = all_orders.filter(status='processing').count()
    shipped_orders = all_orders.filter(status='shipped').count()
    delivered_orders = all_orders.filter(status='delivered').count()
    cancelled_orders = all_orders.filter(status='cancelled').count()

    # Top selling products with revenue
    top_products = products.annotate(
        dashboard_order_count=Count('orderitem__order', distinct=True),
        dashboard_total_sold=Sum('orderitem__quantity'),
        dashboard_total_revenue=Sum('orderitem__price_at_purchase')
    ).order_by('-dashboard_total_revenue')[:5]

    # Low stock products
    low_stock_products = products.filter(stock__lte=15, stock__gt=0).count()

    # Categories for product forms
    categories = Category.objects.all()

    # Get recent activities from database
    recent_activity = seller.activities.filter(is_cleared=False)[:5]

    # Sales chart data (last 6 months)
    sales_data = []
    for i in range(6):
        month_date = datetime.now() - timedelta(days=30*i)
        month_sales = all_orders.filter(
            payment_status='paid',
            created_at__month=month_date.month,
            created_at__year=month_date.year
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        sales_data.append({
            'month': month_date.strftime('%b'),
            'sales': float(month_sales)
        })
    sales_data.reverse()

    # Get notifications from database
    notifications = seller.notifications.filter(is_read=False)[:10]

    # Gift Boxes Data
    # Get active campaigns the seller is participating in
    seller_giftbox_participations = SellerGiftBoxParticipation.objects.filter(
        seller=seller,
        campaign__is_active=True
    ).select_related('campaign')
    
    active_campaigns_count = seller_giftbox_participations.count()
    
    # Get gift box orders for this seller
    giftbox_orders = GiftBoxOrder.objects.filter(seller=seller).select_related('buyer', 'campaign')
    giftbox_orders_count = giftbox_orders.count()
    
    # Calculate gift box revenue
    giftbox_revenue = giftbox_orders.filter(
        status__in=['packed', 'shipped', 'delivered']
    ).aggregate(
        total=Sum('campaign__price')
    )['total'] or 0
    
    # Get pending gift box orders
    pending_giftbox_orders = giftbox_orders.filter(status='pending').count()
    
    # Get recent gift box orders
    recent_giftbox_orders = giftbox_orders.order_by('-created_at')[:10]

    # Promotions Data
    # Get active promotions for this seller
    active_promotions = Promotion.objects.filter(
        seller=seller,
        is_active=True,
        valid_until__gte=timezone.now()
    ).order_by('-created_at')
    
    active_promotions_count = active_promotions.count()
    
    # Calculate total promotion orders from used_count (include all promotions, not just active ones)
    all_promotions = Promotion.objects.filter(seller=seller)
    promotion_orders_count = sum(promotion.used_count for promotion in all_promotions)
    
    # Calculate estimated promotion revenue (this is a simplified approach)
    promotion_revenue = active_promotions.aggregate(
        total=Sum('discount_value')
    )['total'] or 0
    
    # Calculate conversion rate (estimated)
    total_orders_for_conversion = all_orders.count()
    promotion_conversion_rate = 0
    if total_orders_for_conversion > 0:
        promotion_conversion_rate = round((promotion_orders_count / total_orders_for_conversion) * 100, 1)

    context = {
        'seller': seller,
        'recent_orders': recent_orders,
        'products': products,
        'total_earnings': total_earnings,
        'total_orders': total_orders,
        'total_products': total_products,
        'monthly_sales': monthly_sales,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        'cod_revenue': cod_revenue,
        'stripe_revenue': stripe_revenue,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'top_products': top_products,
        'low_stock_products': low_stock_products,
        'recent_activity': recent_activity,
        'sales_data': sales_data,
        'categories': categories,
        'notifications': notifications,
        # Gift Boxes data
        'active_campaigns_count': active_campaigns_count,
        'giftbox_orders_count': giftbox_orders_count,
        'giftbox_revenue': giftbox_revenue,
        'pending_giftbox_orders': pending_giftbox_orders,
        'recent_giftbox_orders': recent_giftbox_orders,
        # Promotions data
        'active_promotions_count': active_promotions_count,
        'promotion_orders_count': promotion_orders_count,
        'promotion_revenue': promotion_revenue,
        'promotion_conversion_rate': promotion_conversion_rate,
        'active_promotions': active_promotions,
    }
    return render(request, 'seller/seller_dashboard.html', context)

@login_required
def seller_orders_list(request):
    """Detailed orders list for seller"""
    seller = get_object_or_404(Seller, user=request.user)
    orders = Order.objects.filter(seller=seller).select_related('buyer').order_by('-created_at')
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(
            Q(id__icontains=search_query) |
            Q(buyer__name__icontains=search_query) |
            Q(buyer__email__icontains=search_query)
        )
    
    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # Apply payment status filter
    payment_status_filter = request.GET.get('payment_status', '')
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'seller': seller,
        'orders': page_obj,
        'total_orders': orders.count(),
    }
    
    # Check if it's an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Return only the orders table and pagination content
        return render(request, 'seller/orders_list_ajax.html', context)
    
    return render(request, 'seller/orders_list.html', context)

@login_required
def seller_products_list(request):
    """Detailed products list for seller"""
    seller = get_object_or_404(Seller, user=request.user)
    products = Product.objects.filter(seller=seller).order_by('-created_at')
    
    # Add search and filter
    search_query = request.GET.get('search', '')
    if search_query:
        products = products.filter(name__icontains=search_query)
    
    category_filter = request.GET.get('category', '')
    if category_filter:
        products = products.filter(category_id=category_filter)
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Category.objects.all()
    
    context = {
        'seller': seller,
        'products': page_obj,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
    }
    return render(request, 'seller/products_list.html', context)

@login_required
def seller_add_product(request):
    """Add new product form"""
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.seller = seller
            product.save()
            
            # Handle multiple images
            images = request.FILES.getlist('images')
            for image in images:
                ProductImage.objects.create(product=product, image=image)
            
            messages.success(request, 'Product added successfully!')
            return redirect('seller:products_list')
    else:
        form = ProductForm()
    
    categories = Category.objects.all()
    
    context = {
        'seller': seller,
        'form': form,
        'categories': categories,
    }
    return render(request, 'seller/product_form.html', context)

@login_required
def seller_edit_product(request, product_id):
    """Edit existing product"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            
            # Handle new images
            images = request.FILES.getlist('images')
            for image in images:
                ProductImage.objects.create(product=product, image=image)
            
            messages.success(request, 'Product updated successfully!')
            return redirect('seller:products_list')
    else:
        form = ProductForm(instance=product)
    
    context = {
        'seller': seller,
        'product': product,
        'form': form,
    }
    return render(request, 'seller/edit_product.html', context)

@login_required
def seller_delete_product(request, product_id):
    """Delete product"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('seller:products_list')
    
    context = {
        'seller': seller,
        'product': product,
    }
    return render(request, 'seller/product_delete.html', context)

@login_required
def seller_reports(request):
    """Sales reports and analytics"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Get orders for this seller
    orders = Order.objects.filter(seller=seller, payment_status='paid')
    
    # Calculate various metrics
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Revenue by period
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    year_ago = today - timedelta(days=365)
    
    weekly_revenue = orders.filter(created_at__date__gte=week_ago).aggregate(
        total=Sum('total_amount'))['total'] or 0
    
    monthly_revenue = orders.filter(created_at__date__gte=month_ago).aggregate(
        total=Sum('total_amount'))['total'] or 0
    
    yearly_revenue = orders.filter(created_at__date__gte=year_ago).aggregate(
        total=Sum('total_amount'))['total'] or 0
    
    # Monthly sales data for chart
    monthly_data = []
    for i in range(6):
        month_start = today.replace(day=1) - timedelta(days=30*i)
        month_end = month_start.replace(day=28) + timedelta(days=4)
        month_end = month_end.replace(day=1) - timedelta(days=1)
        
        month_revenue = orders.filter(
            created_at__date__gte=month_start,
            created_at__date__lte=month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        monthly_data.append({
            'month': month_start.strftime('%b'),
            'revenue': month_revenue
        })
    
    monthly_data.reverse()  # Show oldest to newest
    
    # Top selling products with actual sales data
    top_products = Product.objects.filter(seller=seller).annotate(
        total_sold=Sum('orderitem__quantity'),
        total_revenue=Sum('orderitem__price_at_purchase' * 'orderitem__quantity')
    ).order_by('-total_revenue')[:5]
    
    context = {
        'seller': seller,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'yearly_revenue': yearly_revenue,
        'monthly_data': monthly_data,
        'top_products': top_products,
        'total_orders': orders.count(),
    }
    return render(request, 'seller/reports.html', context)

@login_required
def seller_edit_profile(request):
    """Edit seller profile"""
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        form = SellerUpdateForm(request.POST, request.FILES, instance=seller)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('seller:dashboard')
    else:
        form = SellerUpdateForm(instance=seller)
    
    context = {
        'seller': seller,
        'form': form,
    }
    return render(request, 'seller/edit_profile.html', context)

@login_required
@require_POST
def promotion_create(request):
    """Create a new promotion"""
    seller = get_object_or_404(Seller, user=request.user)
    
    try:
        # Get form data
        name = request.POST.get('name')
        promotion_type = request.POST.get('promotion_type')
        discount_value = request.POST.get('discount_value')
        min_order_amount = request.POST.get('min_order_amount', 0)
        max_discount_amount = request.POST.get('max_discount_amount')
        usage_limit = request.POST.get('usage_limit')
        valid_from_str = request.POST.get('valid_from')
        valid_until_str = request.POST.get('valid_until')
        description = request.POST.get('description', '')
        # Check if the checkbox is checked (sends 'on' when checked, nothing when unchecked)
        is_active = request.POST.get('is_active') == 'on'
        product_ids = request.POST.getlist('products')
        
        # Validate required fields
        if not name or not promotion_type or not discount_value or not valid_from_str or not valid_until_str:
            return JsonResponse({'success': False, 'error': 'All required fields must be filled'})
        
        # Parse dates
        try:
            valid_from = timezone.datetime.fromisoformat(valid_from_str.replace('Z', '+00:00'))
            valid_until = timezone.datetime.fromisoformat(valid_until_str.replace('Z', '+00:00'))
            
            # Make timezone-aware if needed
            if timezone.is_naive(valid_from):
                valid_from = timezone.make_aware(valid_from)
            if timezone.is_naive(valid_until):
                valid_until = timezone.make_aware(valid_until)
                
        except ValueError as e:
            return JsonResponse({'success': False, 'error': f'Invalid date format: {str(e)}'})
        
        # Validate discount value
        try:
            discount_value = float(discount_value)
            if discount_value <= 0:
                return JsonResponse({'success': False, 'error': 'Discount value must be greater than 0'})
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid discount value'})
        
        # Validate dates
        if valid_until <= valid_from:
            return JsonResponse({'success': False, 'error': 'Valid Until date must be after Valid From date'})
        
        # Create the promotion
        promotion = Promotion.objects.create(
            seller=seller,
            name=name,
            description=description,
            promotion_type=promotion_type,
            discount_value=discount_value,
            min_order_amount=float(min_order_amount) if min_order_amount else 0,
            max_discount_amount=float(max_discount_amount) if max_discount_amount else None,
            valid_from=valid_from,
            valid_until=valid_until,
            usage_limit=int(usage_limit) if usage_limit else None,
            is_active=is_active
        )
        
        # Handle product selection
        if product_ids:
            # Get selected products that belong to this seller
            selected_products = Product.objects.filter(id__in=product_ids, seller=seller)
            promotion.products.set(selected_products)
        
        # Create activity for the promotion
        Activity.objects.create(
            seller=seller,
            type='product_updated',
            title=f'Promotion Created: {name}',
            description=f'Created promotion "{name}" with {discount_value} discount'
        )
        
        return JsonResponse({
            'success': True, 
            'message': f'Promotion "{name}" created successfully!'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required
def seller_create_promotion(request):
    """Create promotion for products"""
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            promotion_type = request.POST.get('promotion_type')
            discount_value = request.POST.get('discount_value')
            min_order_amount = request.POST.get('min_order_amount', 0)
            max_discount_amount = request.POST.get('max_discount_amount')
            usage_limit = request.POST.get('usage_limit')
            valid_from_str = request.POST.get('valid_from')
            valid_until_str = request.POST.get('valid_until')
            description = request.POST.get('description', '')
            # Check if the checkbox is checked (sends 'on' when checked, nothing when unchecked)
            is_active = request.POST.get('is_active') == 'on'
            print(f"DEBUG: is_active checkbox value: '{request.POST.get('is_active')}'")
            print(f"DEBUG: is_active boolean: {is_active}")
            
            # Validate required fields
            if not name or not promotion_type or not discount_value or not valid_from_str or not valid_until_str:
                messages.error(request, 'All required fields must be filled.')
                return redirect('sellers:create_promotion')
            
            # Parse dates
            try:
                valid_from = timezone.datetime.fromisoformat(valid_from_str.replace('Z', '+00:00'))
                valid_until = timezone.datetime.fromisoformat(valid_until_str.replace('Z', '+00:00'))
                
                # Make timezone-aware if needed
                if timezone.is_naive(valid_from):
                    valid_from = timezone.make_aware(valid_from)
                if timezone.is_naive(valid_until):
                    valid_until = timezone.make_aware(valid_until)
                
                # Ensure valid_from is not in the future (set to current time if it is)
                now = timezone.now()
                if valid_from > now:
                    print(f"WARNING: valid_from ({valid_from}) is in the future, setting to current time ({now})")
                    valid_from = now
                    
            except ValueError as e:
                messages.error(request, f'Invalid date format: {str(e)}')
                return redirect('sellers:create_promotion')
            
            # Validate discount value
            try:
                discount_value = float(discount_value)
                if discount_value <= 0:
                    messages.error(request, 'Discount value must be greater than 0.')
                    return redirect('sellers:create_promotion')
            except ValueError:
                messages.error(request, 'Invalid discount value.')
                return redirect('sellers:create_promotion')
            
            # Validate dates
            if valid_until <= valid_from:
                messages.error(request, 'Valid Until date must be after Valid From date.')
                return redirect('sellers:create_promotion')
            
            # Create the promotion
            print(f"DEBUG: Creating promotion with valid_from: {valid_from}, valid_until: {valid_until}")
            promotion = Promotion.objects.create(
                seller=seller,
                name=name,
                description=description,
                promotion_type=promotion_type,
                discount_value=discount_value,
                min_order_amount=float(min_order_amount) if min_order_amount else 0,
                max_discount_amount=float(max_discount_amount) if max_discount_amount else None,
                valid_from=valid_from,
                valid_until=valid_until,
                usage_limit=int(usage_limit) if usage_limit else None,
                is_active=is_active
            )
            print(f"DEBUG: Created promotion ID: {promotion.id}, is_valid: {promotion.is_valid}")
            
            # Handle product selection
            product_ids = request.POST.getlist('products')
            if product_ids:
                # Get selected products that belong to this seller
                selected_products = Product.objects.filter(id__in=product_ids, seller=seller)
                promotion.products.set(selected_products)
            
            # Create activity for the promotion
            Activity.objects.create(
                seller=seller,
                type='product_updated',
                title=f'Promotion Created: {name}',
                description=f'Created promotion "{name}" with {discount_value} discount'
            )
            
            messages.success(request, f'Promotion "{name}" created successfully!')
            return redirect('sellers:promotions_list')
            
        except Exception as e:
            messages.error(request, f'Error creating promotion: {str(e)}')
            return redirect('sellers:create_promotion')
    
    products = Product.objects.filter(seller=seller)
    
    context = {
        'seller': seller,
        'products': products,
    }
    return render(request, 'seller/create_promotion.html', context)

@login_required
@require_POST
def export_data(request):
    """Export data in various formats"""
    seller = get_object_or_404(Seller, user=request.user)
    
    try:
        print(f"Export request received for seller: {seller.id}")
        data = json.loads(request.body)
        export_type = data.get('export_type')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        format_type = data.get('format')
        
        print(f"Export parameters: type={export_type}, format={format_type}, start={start_date}, end={end_date}")
        
        if format_type == 'csv':
            print("Exporting as CSV")
            return export_csv(request, seller, export_type, start_date, end_date)
        elif format_type == 'excel':
            print("Exporting as Excel")
            return export_excel(request, seller, export_type, start_date, end_date)
        elif format_type == 'pdf':
            print("Exporting as PDF")
            return export_pdf(request, seller, export_type, start_date, end_date)
        else:
            print(f"Unsupported format: {format_type}")
            return JsonResponse({'success': False, 'error': 'Format not supported'})
    except Exception as e:
        print(f"Export error: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)})

def export_csv(request, seller, export_type, start_date, end_date):
    """Export data as CSV"""
    try:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{export_type}_export_{timezone.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        
        if export_type == 'orders':
            writer.writerow(['Order ID', 'Customer Name', 'Customer Email', 'Total Amount', 'Status', 'Payment Status', 'Order Type', 'Date', 'Delivery Address'])
            orders = Order.objects.filter(seller=seller).select_related('buyer')
            
            # Fix date filtering - only filter if both dates are provided
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                writer.writerow([
                    order.id,
                    order.buyer.name if hasattr(order.buyer, 'name') else order.buyer.user.username,
                    order.buyer.email,
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.delivery_address or 'N/A'
                ])
        
        elif export_type == 'products':
            writer.writerow(['Product ID', 'Name', 'Category', 'Base Price', 'Final Price', 'Stock', 'Condition', 'Status', 'Created Date'])
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                # Use a simple status based on stock availability instead of is_active
                status = 'Active' if product.stock > 0 else 'Out of Stock'
                writer.writerow([
                    product.id,
                    product.name,
                    product.category.name if product.category else 'Uncategorized',
                    product.base_price,
                    product.final_price,
                    product.stock,
                    product.get_condition_display(),
                    status,
                    product.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'sales':
            writer.writerow(['Order ID', 'Product Name', 'Quantity', 'Price at Purchase', 'Total', 'Order Date', 'Status'])
            order_items = OrderItem.objects.filter(order__seller=seller).select_related('order', 'product')
            
            # Fix date filtering for sales
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                order_items = order_items.filter(order__created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for item in order_items:
                writer.writerow([
                    item.order.id,
                    item.product.name,
                    item.quantity,
                    item.price_at_purchase,
                    item.price_at_purchase * item.quantity,
                    item.order.created_at.strftime('%Y-%m-%d %H:%M'),
                    item.order.get_status_display()
                ])
        
        elif export_type == 'giftboxes':
            writer.writerow(['Gift Box Order ID', 'Campaign Name', 'Buyer Name', 'Buyer Email', 'Status', 'Price', 'Created Date', 'Buyer Message'])
            from buyer.models import GiftBoxOrder
            giftbox_orders = GiftBoxOrder.objects.filter(seller=seller).select_related('buyer', 'campaign')
            
            # Fix date filtering for gift boxes
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                writer.writerow([
                    order.id,
                    order.campaign.name,
                    order.buyer.name,
                    order.buyer.email,
                    order.get_status_display(),
                    order.campaign.price,
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.buyer_message or 'N/A'
                ])
        
        elif export_type == 'promotions':
            writer.writerow(['Promotion ID', 'Name', 'Type', 'Discount Value', 'Min Order Amount', 'Valid From', 'Valid Until', 'Status', 'Used Count', 'Created Date'])
            from seller.models import Promotion
            promotions = Promotion.objects.filter(seller=seller)
            
            # Fix date filtering for promotions
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotions = promotions.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for promotion in promotions:
                writer.writerow([
                    promotion.id,
                    promotion.name,
                    promotion.get_promotion_type_display(),
                    promotion.discount_value,
                    promotion.min_order_amount,
                    promotion.valid_from.strftime('%Y-%m-%d %H:%M'),
                    promotion.valid_until.strftime('%Y-%m-%d %H:%M'),
                    'Active' if promotion.is_active else 'Inactive',
                    promotion.used_count,
                    promotion.created_at.strftime('%Y-%m-%d %H:%M')
                ])
        
        elif export_type == 'inventory':
            writer.writerow(['Product ID', 'Name', 'Category', 'Current Stock', 'Low Stock Alert', 'Total Sold', 'Revenue Generated', 'Last Updated'])
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                low_stock = 'Yes' if product.stock <= 15 else 'No'
                
                # Calculate totals manually to avoid type issues
                order_items = OrderItem.objects.filter(product=product)
                total_sold = sum(item.quantity for item in order_items)
                revenue_generated = sum(item.price_at_purchase * item.quantity for item in order_items)
                
                writer.writerow([
                    product.id,
                    product.name,
                    product.category.name if product.category else 'Uncategorized',
                    product.stock,
                    low_stock,
                    total_sold,
                    revenue_generated,
                    product.created_at.strftime('%Y-%m-%d %H:%M')
                ])
        
        return response
    except Exception as e:
        print(f"CSV export error: {str(e)}")
        return JsonResponse({'success': False, 'error': f'CSV export failed: {str(e)}'})

def export_excel(request, seller, export_type, start_date, end_date):
    """Export data as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return JsonResponse({'success': False, 'error': 'Excel export requires openpyxl package'})
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = export_type.title()
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        if export_type == 'orders':
            headers = ['Order ID', 'Customer Name', 'Customer Email', 'Total Amount', 'Status', 'Payment Status', 'Order Type', 'Date', 'Delivery Address']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            orders = Order.objects.filter(seller=seller).select_related('buyer')
            
            # Fix date filtering - only filter if both dates are provided
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                ws.append([
                    order.id,
                    order.buyer.name if hasattr(order.buyer, 'name') else order.buyer.user.username,
                    order.buyer.email,
                    order.total_amount,
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.delivery_address or 'N/A'
                ])
        
        elif export_type == 'products':
            headers = ['Product ID', 'Name', 'Category', 'Base Price', 'Final Price', 'Stock', 'Condition', 'Status', 'Created Date']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                # Use a simple status based on stock availability instead of is_active
                status = 'Active' if product.stock > 0 else 'Out of Stock'
                ws.append([
                    product.id,
                    product.name,
                    product.category.name if product.category else 'Uncategorized',
                    product.base_price,
                    product.final_price,
                    product.stock,
                    product.get_condition_display(),
                    status,
                    product.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'sales':
            headers = ['Order ID', 'Product Name', 'Quantity', 'Price at Purchase', 'Total', 'Order Date', 'Status']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            order_items = OrderItem.objects.filter(order__seller=seller).select_related('order', 'product')
            
            # Fix date filtering for sales
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                order_items = order_items.filter(order__created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for item in order_items:
                ws.append([
                    item.order.id,
                    item.product.name,
                    item.quantity,
                    item.price_at_purchase,
                    item.price_at_purchase * item.quantity,
                    item.order.created_at.strftime('%Y-%m-%d'),
                    item.order.get_status_display()
                ])
        
        elif export_type == 'inventory':
            headers = ['Product ID', 'Name', 'Category', 'Current Stock', 'Low Stock Alert', 'Total Sold', 'Revenue Generated', 'Last Updated']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                low_stock = 'Yes' if product.stock <= 15 else 'No'
                
                # Calculate totals manually to avoid type issues
                order_items = OrderItem.objects.filter(product=product)
                total_sold = sum(item.quantity for item in order_items)
                revenue_generated = sum(item.price_at_purchase * item.quantity for item in order_items)
                
                ws.append([
                    product.id,
                    product.name,
                    product.category.name if product.category else 'Uncategorized',
                    product.stock,
                    low_stock,
                    total_sold,
                    revenue_generated,
                    product.created_at.strftime('%Y-%m-%d %H:%M')
                ])
        
        elif export_type == 'giftboxes':
            headers = ['Gift Box Order ID', 'Campaign Name', 'Buyer Name', 'Buyer Email', 'Status', 'Price', 'Created Date', 'Buyer Message']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            from buyer.models import GiftBoxOrder
            giftbox_orders = GiftBoxOrder.objects.filter(seller=seller).select_related('buyer', 'campaign')
            
            # Fix date filtering for gift boxes
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                ws.append([
                    order.id,
                    order.campaign.name,
                    order.buyer.name,
                    order.buyer.email,
                    order.get_status_display(),
                    order.campaign.price,
                    order.created_at.strftime('%Y-%m-%d %H:%M'),
                    order.buyer_message or 'N/A'
                ])
        
        elif export_type == 'promotions':
            headers = ['Promotion ID', 'Name', 'Type', 'Discount Value', 'Min Order Amount', 'Valid From', 'Valid Until', 'Status', 'Used Count', 'Created Date']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            from seller.models import Promotion
            promotions = Promotion.objects.filter(seller=seller)
            
            # Fix date filtering for promotions
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotions = promotions.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for promotion in promotions:
                ws.append([
                    promotion.id,
                    promotion.name,
                    promotion.get_promotion_type_display(),
                    promotion.discount_value,
                    promotion.min_order_amount,
                    promotion.valid_from.strftime('%Y-%m-%d %H:%M'),
                    promotion.valid_until.strftime('%Y-%m-%d %H:%M'),
                    'Active' if promotion.is_active else 'Inactive',
                    promotion.used_count,
                    promotion.created_at.strftime('%Y-%m-%d %H:%M')
                ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{export_type}_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Excel export failed: {str(e)}'})

def export_pdf(request, seller, export_type, start_date, end_date):
    """Export data as PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return JsonResponse({'success': False, 'error': 'PDF export requires reportlab package'})
    
    try:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{export_type}_export_{timezone.now().strftime("%Y%m%d")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        title = Paragraph(f"{export_type.title()} Report - {seller.shop_name}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        if export_type == 'orders':
            headers = ['Order ID', 'Customer', 'Total', 'Status', 'Payment', 'Type', 'Date']
            data = [headers]
            
            orders = Order.objects.filter(seller=seller).select_related('buyer')
            
            # Fix date filtering - only filter if both dates are provided
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                orders = orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in orders:
                data.append([
                    str(order.id),
                    order.buyer.name if hasattr(order.buyer, 'name') else order.buyer.user.username,
                    f"${order.total_amount}",
                    order.get_status_display(),
                    order.get_payment_status_display(),
                    order.order_type.upper(),
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'products':
            headers = ['Product ID', 'Name', 'Category', 'Price', 'Stock', 'Status']
            data = [headers]
            
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                # Use a simple status based on stock availability instead of is_active
                status = 'Active' if product.stock > 0 else 'Out of Stock'
                data.append([
                    str(product.id),
                    product.name,
                    product.category.name if product.category else 'Uncategorized',
                    f"${product.final_price}",
                    str(product.stock),
                    status
                ])
        
        elif export_type == 'sales':
            headers = ['Order ID', 'Product', 'Quantity', 'Price', 'Total', 'Date']
            data = [headers]
            
            order_items = OrderItem.objects.filter(order__seller=seller).select_related('order', 'product')
            
            # Fix date filtering for sales
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                order_items = order_items.filter(order__created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for item in order_items:
                data.append([
                    str(item.order.id),
                    item.product.name,
                    str(item.quantity),
                    f"${item.price_at_purchase}",
                    f"${item.price_at_purchase * item.quantity}",
                    item.order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'inventory':
            headers = ['Product ID', 'Name', 'Stock', 'Low Stock', 'Total Sold', 'Revenue']
            data = [headers]
            
            products = Product.objects.filter(seller=seller).select_related('category')
            
            for product in products:
                low_stock = 'Yes' if product.stock <= 15 else 'No'
                
                # Calculate totals manually to avoid type issues
                order_items = OrderItem.objects.filter(product=product)
                total_sold = sum(item.quantity for item in order_items)
                revenue_generated = sum(item.price_at_purchase * item.quantity for item in order_items)
                
                data.append([
                    str(product.id),
                    product.name,
                    str(product.stock),
                    low_stock,
                    str(total_sold),
                    f"${revenue_generated}"
                ])
        
        elif export_type == 'giftboxes':
            headers = ['Gift Box ID', 'Campaign', 'Buyer', 'Status', 'Price', 'Date']
            data = [headers]
            
            from buyer.models import GiftBoxOrder
            giftbox_orders = GiftBoxOrder.objects.filter(seller=seller).select_related('buyer', 'campaign')
            
            # Fix date filtering for gift boxes
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                giftbox_orders = giftbox_orders.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for order in giftbox_orders:
                data.append([
                    str(order.id),
                    order.campaign.name,
                    order.buyer.name,
                    order.get_status_display(),
                    f"${order.campaign.price}",
                    order.created_at.strftime('%Y-%m-%d')
                ])
        
        elif export_type == 'promotions':
            headers = ['Promotion ID', 'Name', 'Type', 'Discount', 'Status', 'Used Count', 'Created Date']
            data = [headers]
            
            from seller.models import Promotion
            promotions = Promotion.objects.filter(seller=seller)
            
            # Fix date filtering for promotions
            if start_date and end_date and start_date.strip() and end_date.strip():
                from datetime import datetime
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                promotions = promotions.filter(created_at__date__range=[start_dt.date(), end_dt.date()])
            
            for promotion in promotions:
                data.append([
                    str(promotion.id),
                    promotion.name,
                    promotion.get_promotion_type_display(),
                    f"{promotion.discount_value}%",
                    'Active' if promotion.is_active else 'Inactive',
                    str(promotion.used_count),
                    promotion.created_at.strftime('%Y-%m-%d')
                ])
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        return response
        
    except Exception as e:
        print(f"PDF export error: {str(e)}")
        return JsonResponse({'success': False, 'error': f'PDF export failed: {str(e)}'})

@login_required
@require_GET
def check_new_orders(request):
    """Check for new orders (for real-time notifications)"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Check for orders created in the last 5 minutes
    recent_orders = Order.objects.filter(
        seller=seller,
        created_at__gte=datetime.now() - timedelta(minutes=5)
    ).count()
    
    return JsonResponse({'new_orders': recent_orders})

@login_required
@require_GET
def check_low_stock(request):
    """Check for products with low stock"""
    seller = get_object_or_404(Seller, user=request.user)
    
    low_stock_products = Product.objects.filter(
        seller=seller,
        stock__lte=15,
        stock__gt=0
    ).count()
    
    return JsonResponse({'low_stock_products': low_stock_products})

def calculate_seller_stats(seller):
    """Calculate seller statistics for dashboard"""
    from datetime import datetime, timedelta
    
    # Get current month
    now = datetime.now()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # COD revenue (paid COD orders that are delivered)
    cod_revenue = Order.objects.filter(
        seller=seller,
        order_type='cod',
        payment_status='paid',
        status='delivered'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Stripe revenue (paid Stripe orders)
    stripe_revenue = Order.objects.filter(
        seller=seller,
        order_type='stripe',
        payment_status='paid'
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Total earnings = COD revenue + Stripe revenue
    total_earnings = cod_revenue + stripe_revenue
    
    # Monthly sales (current month) - COD (paid + delivered) + Stripe (paid)
    monthly_cod_revenue = Order.objects.filter(
        seller=seller,
        order_type='cod',
        payment_status='paid',
        status='delivered',
        created_at__gte=start_of_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_stripe_revenue = Order.objects.filter(
        seller=seller,
        order_type='stripe',
        payment_status='paid',
        created_at__gte=start_of_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_sales = monthly_cod_revenue + monthly_stripe_revenue
    
    return {
        'total_earnings': total_earnings,
        'monthly_sales': monthly_sales,
        'cod_revenue': cod_revenue,
        'stripe_revenue': stripe_revenue
    }

@login_required
@require_POST
def mark_notification_as_read(request):
    """Mark a specific notification as read"""
    try:
        notification_id = request.POST.get('notification_id')
        seller = get_object_or_404(Seller, user=request.user)
        
        notification = get_object_or_404(Notification, id=notification_id, seller=seller)
        notification.is_read = True
        notification.save()
        
        return JsonResponse({'success': True, 'message': 'Notification marked as read'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def mark_all_notifications_as_read(request):
    """Mark all notifications as read for a seller"""
    try:
        seller = get_object_or_404(Seller, user=request.user)
        
        # Mark all unread notifications as read
        updated_count = seller.notifications.filter(is_read=False).update(is_read=True)
        
        return JsonResponse({
            'success': True, 
            'message': f'{updated_count} notifications marked as read'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def clear_all_notifications(request):
    """Delete all notifications for a seller"""
    try:
        seller = get_object_or_404(Seller, user=request.user)
        
        # Delete all notifications
        deleted_count = seller.notifications.count()
        seller.notifications.all().delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'{deleted_count} notifications cleared'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def clear_all_activity(request):
    """Clear all activities for a seller"""
    try:
        seller = get_object_or_404(Seller, user=request.user)
        
        # Mark all activities as cleared
        updated_count = seller.activities.filter(is_cleared=False).update(is_cleared=True)
        
        return JsonResponse({
            'success': True, 
            'message': f'{updated_count} activities cleared'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

@login_required
@require_POST
def mark_activity_as_cleared(request):
    """Mark a specific activity as cleared"""
    try:
        activity_id = request.POST.get('activity_id')
        seller = get_object_or_404(Seller, user=request.user)
        
        activity = get_object_or_404(Activity, id=activity_id, seller=seller)
        activity.is_cleared = True
        activity.save()
        
        return JsonResponse({'success': True, 'message': 'Activity marked as cleared'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)

# ========== HOT OFFERS FEATURE ==========

def hot_offers_view(request):
    """Display all promotions across the platform"""
    from django.db.models import Q
    from django.utils import timezone
    
    # Get only non-expired promotions
    now = timezone.now()
    promotions = Promotion.objects.filter(
        valid_until__gt=now,  # Only promotions that haven't expired
        is_active=True  # Only active promotions
    ).select_related('seller').prefetch_related('products', 'categories').order_by('-created_at')
    
    # Get all categories for filtering
    categories = Category.objects.all()
    
    # Handle search and filtering
    search_query = request.GET.get('search', '')
    category_filter = request.GET.getlist('category')
    
    if search_query:
        promotions = promotions.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(seller__shop_name__icontains=search_query) |
            Q(products__name__icontains=search_query)
        ).distinct()
    
    if category_filter:
        promotions = promotions.filter(
            Q(categories__id__in=category_filter) |
            Q(products__category__id__in=category_filter)
        ).distinct()
    
    # Get user info for context
    user = request.user if request.user.is_authenticated else None
    seller = None
    buyer = None
    user_type = None
    
    if user:
        try:
            seller = Seller.objects.get(user=user)
            user_type = 'seller'
        except Seller.DoesNotExist:
            try:
                buyer = Buyer.objects.get(email=user.email)
                user_type = 'buyer'
            except Buyer.DoesNotExist:
                pass
    
    context = {
        'promotions': promotions,
        'categories': categories,
        'search_query': search_query,
        'category_filter': category_filter,
        'user': user,
        'seller': seller,
        'buyer': buyer,
        'user_type': user_type,
    }
    
    return render(request, 'seller/hot_offers.html', context)

def promotion_detail_view(request, promotion_id):
    """Display detailed information about a specific promotion"""
    promotion = get_object_or_404(Promotion, id=promotion_id)
    
    # Get user info for context
    user = request.user if request.user.is_authenticated else None
    seller = None
    buyer = None
    user_type = None
    
    if user:
        try:
            seller = Seller.objects.get(user=user)
            user_type = 'seller'
        except Seller.DoesNotExist:
            try:
                buyer = Buyer.objects.get(email=user.email)
                user_type = 'buyer'
            except Buyer.DoesNotExist:
                pass
    
    # Get products in this promotion
    promotion_products = promotion.products.all().prefetch_related('images', 'category')
    
    # Calculate time remaining
    from django.utils import timezone
    now = timezone.now()
    time_remaining = promotion.valid_until - now
    
    context = {
        'promotion': promotion,
        'promotion_products': promotion_products,
        'time_remaining': time_remaining,
        'user': user,
        'seller': seller,
        'buyer': buyer,
        'user_type': user_type,
    }
    
    return render(request, 'seller/promotion_detail.html', context)

def search_promotions_ajax(request):
    """AJAX endpoint for searching promotions"""
    from django.db.models import Q
    from django.utils import timezone
    
    search_query = request.GET.get('q', '')
    category_filter = request.GET.getlist('category')
    
    # Get active promotions
    promotions = Promotion.objects.filter(
        is_active=True,
        valid_from__lte=timezone.now(),
        valid_until__gte=timezone.now()
    ).select_related('seller').prefetch_related('products', 'categories')
    
    # Apply search filter
    if search_query:
        promotions = promotions.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(seller__shop_name__icontains=search_query) |
            Q(products__name__icontains=search_query)
        ).distinct()
    
    # Apply category filter
    if category_filter:
        promotions = promotions.filter(
            Q(categories__id__in=category_filter) |
            Q(products__category__id__in=category_filter)
        ).distinct()
    
    # Serialize promotions for JSON response
    promotions_data = []
    for promotion in promotions:
        promotion_data = {
            'id': promotion.id,
            'name': promotion.name,
            'description': promotion.description,
            'promotion_type': promotion.promotion_type,
            'discount_value': promotion.discount_value,
            'seller_name': promotion.seller.shop_name,
            'valid_until': promotion.valid_until.isoformat(),
            'products_count': promotion.products.count(),
            'url': f'/promotion/{promotion.id}/'
        }
        promotions_data.append(promotion_data)
    
    return JsonResponse({'promotions': promotions_data})

def get_active_promotions_for_homepage():
    """Get active promotions for homepage display - 5 most ordered promotions"""
    from django.utils import timezone
    from django.db import models
    
    # Get active promotions that are not expired
    active_promotions = Promotion.objects.filter(
        is_active=True,
        valid_from__lte=timezone.now(),
        valid_until__gte=timezone.now()
    ).select_related('seller').prefetch_related('products')
    
    # Get top 5 most ordered promotions
    most_ordered_promotions = active_promotions.order_by('-used_count', 'valid_until')[:5]
    
    return most_ordered_promotions

@login_required
def seller_export_data(request):
    """Handle export data requests"""
    seller = get_object_or_404(Seller, user=request.user)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            export_type = data.get('export_type')
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            format_type = data.get('format')
            
            if format_type == 'csv':
                return export_csv(request, seller, export_type, start_date, end_date)
            elif format_type == 'excel':
                return export_excel(request, seller, export_type, start_date, end_date)
            elif format_type == 'pdf':
                return export_pdf(request, seller, export_type, start_date, end_date)
            else:
                return JsonResponse({'success': False, 'error': 'Format not supported'})
                
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def test_export(request):
    """Test export functionality"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Test CSV export
    try:
        csv_response = export_csv(request, seller, 'products', None, None)
        print("CSV export test successful")
    except Exception as e:
        print(f"CSV export test failed: {e}")
    
    # Test Excel export
    try:
        excel_response = export_excel(request, seller, 'products', None, None)
        print("Excel export test successful")
    except Exception as e:
        print(f"Excel export test failed: {e}")
    
    # Test PDF export
    try:
        pdf_response = export_pdf(request, seller, 'products', None, None)
        print("PDF export test successful")
    except Exception as e:
        print(f"PDF export test failed: {e}")
    
    return JsonResponse({'success': True, 'message': 'Export tests completed'})

@login_required
def simple_export_test(request):
    """Simple test to verify export works"""
    seller = get_object_or_404(Seller, user=request.user)
    
    try:
        # Create a simple CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="test_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Test', 'Data'])
        writer.writerow(['Product', 'Test Product'])
        writer.writerow(['Price', '100'])
        
        print("Simple export test successful")
        return response
        
    except Exception as e:
        print(f"Simple export test failed: {e}")
        return JsonResponse({'success': False, 'error': str(e)})

# Missing functions that are referenced in URLs
@login_required
def product_edit_data(request, product_id):
    """Get product data for editing"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    context = {
        'product': product,
        'seller': seller,
    }
    return render(request, 'seller/product_edit_data.html', context)

@login_required
@require_POST
def product_update(request, product_id):
    """Update product data"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    try:
        # Update basic fields
        product.name = request.POST.get('name', product.name)
        product.description = request.POST.get('description', product.description)
        product.base_price = float(request.POST.get('base_price', product.base_price))
        product.stock = int(request.POST.get('stock', product.stock))
        product.condition = request.POST.get('condition', product.condition)
        
        # Auto-calculate final price based on discount
        discount = request.POST.get('discount_percentage')
        if discount:
            product.discount_percentage = float(discount)
        else:
            product.discount_percentage = 0
        
        base_price = product.base_price or 0
        discount_percentage = product.discount_percentage or 0
        
        if discount_percentage > 0:
            # Calculate final price with discount
            discount_amount = base_price * (discount_percentage / 100)
            product.final_price = max(0, base_price - discount_amount)
        else:
            # If no discount, final price equals base price
            product.final_price = base_price
        
        product.save()
        
        return JsonResponse({'success': True, 'message': 'Product updated successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def product_delete(request, product_id):
    """Delete product"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    try:
        product_name = product.name
        product.delete()
        return JsonResponse({'success': True, 'message': f'Product "{product_name}" deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_POST
def product_duplicate(request, product_id):
    """Duplicate product"""
    seller = get_object_or_404(Seller, user=request.user)
    product = get_object_or_404(Product, id=product_id, seller=seller)
    
    try:
        # Create a copy of the product
        new_product = Product.objects.create(
            seller=seller,
            category=product.category,
            brand=product.brand,
            name=f"{product.name} (Copy)",
            description=product.description,
            base_price=product.base_price,
            discount_percentage=product.discount_percentage,
            final_price=product.final_price,
            stock=product.stock,
            condition=product.condition,
            rating_avg=product.rating_avg,
            order_count=0
        )
        
        # Copy images
        for image in product.images.all():
            ProductImage.objects.create(
                product=new_product,
                image=image.image,
                is_thumbnail=image.is_thumbnail
            )
        
        # Copy attribute values
        for attr_value in product.attribute_values.all():
            ProductAttributeValue.objects.create(
                product=new_product,
                attribute=attr_value.attribute,
                value=attr_value.value
            )
        
        return JsonResponse({
            'success': True, 
            'message': f'Product "{product.name}" duplicated successfully',
            'new_product_id': new_product.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def order_details(request, order_id):
    """Get order details"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Check if it's a gift box order first
    from buyer.models import GiftBoxOrder
    try:
        giftbox_order = GiftBoxOrder.objects.get(id=order_id, seller=seller)
        # Handle gift box order
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Get buyer information
            buyer_info = {
                'name': getattr(giftbox_order.buyer, 'name', None) or getattr(giftbox_order.buyer.user, 'username', 'Anonymous'),
                'email': getattr(giftbox_order.buyer, 'email', 'N/A'),
                'phone': getattr(giftbox_order.buyer, 'phone', 'N/A'),
            }
            
            # Get payment information (gift box orders use campaign price)
            payment_info = {
                'method': 'GIFT_BOX',
                'transaction_id': 'N/A',
                'status': 'paid',  # Gift boxes are paid upfront
                'payment_time': giftbox_order.created_at.isoformat()
            }
            
            return JsonResponse({
                'id': giftbox_order.id,
                'customer_name': buyer_info['name'],
                'customer_email': buyer_info['email'],
                'order_type': 'giftbox',
                'status': giftbox_order.status,
                'payment_status': 'paid',
                'total_amount': float(giftbox_order.campaign.price),
                'tracking_number': '',  # Gift boxes don't have tracking initially
                'notes': giftbox_order.buyer_message or '',
                'created_at': giftbox_order.created_at.isoformat(),
                'delivery_address': giftbox_order.delivery_address or 'N/A',
                # Add display values
                'get_status_display': giftbox_order.get_status_display(),
                'get_payment_status_display': 'Paid',
                # Add buyer and payment objects
                'buyer': buyer_info,
                'payment': payment_info,
                # Add order items (gift box items are selected by seller)
                'items': []
            })
        
        context = {
            'order': giftbox_order,
            'seller': seller,
        }
        return render(request, 'seller/order_details.html', context)
        
    except GiftBoxOrder.DoesNotExist:
        # Handle regular order
        order = get_object_or_404(Order, id=order_id, seller=seller)
        
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Check if this is a promotion order
            is_promotion_order = order.notes and order.notes.startswith('Promotion:')
            promotion_name = order.notes.replace('Promotion: ', '') if is_promotion_order else ''
            
            # Get order items
            order_items = []
            for item in order.items.all():
                order_items.append({
                    'id': item.id,
                    'product': {
                        'id': item.product.id,
                        'name': item.product.name,
                    },
                    'quantity': item.quantity,
                    'price_at_purchase': float(item.price_at_purchase),
                    'total': float(item.price_at_purchase * item.quantity)
                })
            
            # Get buyer information
            buyer_info = {
                'name': getattr(order.buyer, 'name', None) or getattr(order.buyer.user, 'username', 'Anonymous'),
                'email': getattr(order.buyer, 'email', 'N/A'),
                'phone': getattr(order.buyer, 'phone', 'N/A'),
            }
            
            # Get payment information
            payment_info = {
                'method': 'PROMOTION' if is_promotion_order else order.order_type.upper(),
                'transaction_id': getattr(order, 'stripe_payment_intent_id', 'N/A'),
                'status': order.payment_status,
                'payment_time': order.created_at.isoformat() if order.payment_status == 'paid' else None
            }
            
            return JsonResponse({
            'id': order.id,
            'customer_name': buyer_info['name'],
            'customer_email': buyer_info['email'],
            'order_type': 'promotion' if is_promotion_order else order.order_type,
            'status': order.status,
            'payment_status': order.payment_status,
            'total_amount': float(order.total_amount),
            'tracking_number': order.tracking_number or '',
            'notes': order.notes or '',
            'created_at': order.created_at.isoformat(),
            'delivery_address': order.delivery_address or 'N/A',
            # Add display values
            'get_status_display': order.get_status_display(),
            'get_payment_status_display': order.get_payment_status_display(),
            # Add buyer and payment objects
            'buyer': buyer_info,
            'payment': payment_info,
            # Add order items
            'items': order_items,
            # Add promotion info if applicable
            'is_promotion': is_promotion_order,
            'promotion_name': promotion_name
        })
    
    context = {
        'order': order,
        'seller': seller,
    }
    return render(request, 'seller/order_details.html', context)

@login_required
@require_POST
def order_update_status(request):
    """Update order status"""
    seller = get_object_or_404(Seller, user=request.user)
    
    try:
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        payment_status = request.POST.get('payment_status')
        tracking_number = request.POST.get('tracking_number')
        notes = request.POST.get('notes')
        
        order = get_object_or_404(Order, id=order_id, seller=seller)
        
        # Track what was updated
        updates_made = []
        
        # Update order status
        if new_status and new_status != order.status:
            order.status = new_status
            updates_made.append(f'Status: {new_status}')
        
        # Update payment status if provided
        if payment_status and payment_status != order.payment_status:
            order.payment_status = payment_status
            updates_made.append(f'Payment: {payment_status}')
        
        # Update tracking number if provided
        if tracking_number and tracking_number != order.tracking_number:
            order.tracking_number = tracking_number
            updates_made.append(f'Tracking: {tracking_number}')
        
        # Update notes if provided
        if notes and notes != order.notes:
            order.notes = notes
            updates_made.append('Notes updated')
        
        order.save()
        
        # Create activity for status update
        if updates_made:
            Activity.objects.create(
                seller=seller,
                type='order_updated',
                title=f'Order #{order_id} Updated',
                description=f'Updated: {", ".join(updates_made)}'
            )
        
        # Calculate updated stats for dashboard
        updated_stats = calculate_seller_stats(seller)
        
        return JsonResponse({
            'success': True, 
            'message': f'Order #{order_id} updated successfully',
            'order_type': order.order_type,
            'payment_updated': payment_status and payment_status != order.payment_status,
            'updated_stats': updated_stats
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def test_create_promotion(request):
    """Test function to create a promotion with correct dates"""
    from datetime import timedelta
    
    seller = get_object_or_404(Seller, user=request.user)
    
    # Create a test promotion that expires in 30 days
    now = timezone.now()
    valid_until = now + timedelta(days=30)
    
    promotion = Promotion.objects.create(
        seller=seller,
        name="Test Countdown Promotion",
        description="Testing countdown functionality",
        promotion_type='percentage',
        discount_value=20.0,
        valid_from=now,
        valid_until=valid_until,
        is_active=True
    )
    
    return JsonResponse({
        'success': True,
        'message': f'Test promotion created with valid_until: {valid_until}',
        'promotion_id': promotion.id
    })

@login_required
def giftbox_campaigns_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    from django.utils import timezone
    today = timezone.now().date()
    
    # Get all active campaigns that haven't expired
    campaigns = GiftBoxCampaign.objects.filter(
        is_active=True,
        end_date__gte=today  # Only show campaigns that haven't expired
    ).order_by('-start_date')
    
    participations = SellerGiftBoxParticipation.objects.filter(seller=seller)
    joined_campaign_ids = participations.values_list('campaign_id', flat=True)
    return render(request, 'seller/giftbox_campaigns.html', {
        'campaigns': campaigns,
        'joined_campaign_ids': joined_campaign_ids,
    })

@login_required
def join_giftbox_campaign(request, campaign_id):
    seller = get_object_or_404(Seller, user=request.user)
    from django.utils import timezone
    today = timezone.now().date()
    
    campaign = get_object_or_404(
        GiftBoxCampaign, 
        id=campaign_id, 
        is_active=True,
        end_date__gte=today  # Only allow joining campaigns that haven't expired
    )
    
    participation, created = SellerGiftBoxParticipation.objects.get_or_create(seller=seller, campaign=campaign)
    if created:
        messages.success(request, f'You have joined the campaign: {campaign.name}!')
    else:
        messages.info(request, f'You have already joined this campaign.')
    return redirect(reverse('seller:giftbox_campaigns'))

@login_required
def giftbox_orders_seller_view(request):
    seller = get_object_or_404(Seller, user=request.user)
    orders = GiftBoxOrder.objects.filter(seller=seller).select_related('buyer', 'campaign').order_by('-created_at')
    return render(request, 'seller/giftbox_orders.html', {'orders': orders})

@login_required
def fulfill_giftbox_order_view(request, order_id):
    seller = get_object_or_404(Seller, user=request.user)
    order = get_object_or_404(GiftBoxOrder, id=order_id, seller=seller)
    products = Product.objects.filter(seller=seller)
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_products')
        order.selected_products.set(selected_ids)
        new_status = request.POST.get('status')
        if new_status in dict(GiftBoxOrder.STATUS_CHOICES):
            order.status = new_status
            if new_status == 'delivered':
                order.reveal_contents = True
        order.save()
        messages.success(request, 'Gift box order updated!')
        return redirect('seller:giftbox_orders')
    return render(request, 'seller/fulfill_giftbox_order.html', {
        'order': order,
        'products': products,
        'selected_ids': order.selected_products.values_list('id', flat=True),
    })

@login_required
@require_POST
def delete_promotion(request, promotion_id):
    """Delete a promotion"""
    seller = get_object_or_404(Seller, user=request.user)
    promotion = get_object_or_404(Promotion, id=promotion_id, seller=seller)
    
    promotion_name = promotion.name
    promotion.delete()
    
    messages.success(request, f'Promotion "{promotion_name}" deleted successfully!')
    return JsonResponse({'success': True, 'message': 'Promotion deleted successfully'})

@login_required
def promotion_update_form_view(request, promotion_id):
    """Get promotion update form"""
    seller = get_object_or_404(Seller, user=request.user)
    promotion = get_object_or_404(Promotion, id=promotion_id, seller=seller)
    products = Product.objects.filter(seller=seller)
    
    context = {
        'promotion': promotion,
        'products': products,
        'is_update': True,
        'form_title': 'Update Promotion',
        'submit_text': 'Update Promotion'
    }
    return render(request, 'seller/promotion_update_modal.html', context)

@login_required
def promotion_update_view(request, promotion_id):
    """Update a promotion"""
    seller = get_object_or_404(Seller, user=request.user)
    promotion = get_object_or_404(Promotion, id=promotion_id, seller=seller)
    
    if request.method == 'POST':
        try:
            # Parse form data similar to seller_create_promotion
            name = request.POST.get('name')
            description = request.POST.get('description', '')
            promotion_type = request.POST.get('promotion_type')
            discount_value = request.POST.get('discount_value')
            min_order_amount = request.POST.get('min_order_amount')
            max_discount_amount = request.POST.get('max_discount_amount')
            valid_from_str = request.POST.get('valid_from')
            valid_until_str = request.POST.get('valid_until')
            usage_limit = request.POST.get('usage_limit')
            is_active = request.POST.get('is_active') == 'on'
            
            # Validation
            if not all([name, promotion_type, discount_value, valid_from_str, valid_until_str]):
                return JsonResponse({'success': False, 'error': 'All required fields must be filled'})
            
            # Parse dates
            try:
                valid_from = timezone.datetime.fromisoformat(valid_from_str.replace('Z', '+00:00'))
                valid_until = timezone.datetime.fromisoformat(valid_until_str.replace('Z', '+00:00'))
                
                if timezone.is_naive(valid_from):
                    valid_from = timezone.make_aware(valid_from)
                if timezone.is_naive(valid_until):
                    valid_until = timezone.make_aware(valid_until)
                    
            except ValueError as e:
                return JsonResponse({'success': False, 'error': f'Invalid date format: {str(e)}'})
            
            # Update promotion
            promotion.name = name
            promotion.description = description
            promotion.promotion_type = promotion_type
            promotion.discount_value = float(discount_value)
            promotion.min_order_amount = float(min_order_amount) if min_order_amount else 0
            promotion.max_discount_amount = float(max_discount_amount) if max_discount_amount else None
            promotion.valid_from = valid_from
            promotion.valid_until = valid_until
            promotion.usage_limit = int(usage_limit) if usage_limit else None
            promotion.is_active = is_active
            promotion.save()
            
            # Update products
            selected_products = request.POST.getlist('products')
            promotion.products.clear()
            if selected_products:
                products = Product.objects.filter(id__in=selected_products, seller=seller)
                promotion.products.add(*products)
            
            # Create activity
            Activity.objects.create(
                seller=seller,
                type='product_updated',
                title=f'Promotion Updated: {name}',
                description=f'Promotion "{name}" was updated successfully'
            )
            
            return JsonResponse({
                'success': True,
                'promotion': {
                    'id': promotion.id,
                    'name': promotion.name,
                    'promotion_type': promotion.promotion_type,
                    'discount_value': promotion.discount_value,
                    'is_active': promotion.is_active
                }
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def promotions_list_view(request):
    """List all promotions for the seller"""
    seller = get_object_or_404(Seller, user=request.user)
    
    # Get all promotions for this seller
    promotions = Promotion.objects.filter(seller=seller).order_by('-created_at')
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        promotions = promotions.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter == 'active':
        promotions = promotions.filter(is_active=True, valid_until__gte=timezone.now())
    elif status_filter == 'inactive':
        promotions = promotions.filter(Q(is_active=False) | Q(valid_until__lt=timezone.now()))
    
    # Add pagination
    from django.core.paginator import Paginator
    paginator = Paginator(promotions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'seller': seller,
        'promotions': page_obj,
        'total_promotions': promotions.count(),
    }
    
    return render(request, 'seller/promotions_list.html', context)

def product_listing_view(request):
    """Comprehensive product listing page showing all products from all sellers"""
    # Get all products with related data
    products = Product.objects.select_related('seller', 'category', 'brand').prefetch_related('images').all()
    
    # Get filter parameters
    parent_category_id = request.GET.get('parent_category')
    child_category_id = request.GET.get('child_category')
    brand_id = request.GET.get('brand')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    condition = request.GET.get('condition')
    rating = request.GET.get('rating')
    search_query = request.GET.get('search')
    sort_by = request.GET.get('sort', 'featured')
    view_type = request.GET.get('view', 'grid')
    verified_only = request.GET.get('verified') == 'on'
    
    # Apply filters
    if parent_category_id:
        # Get the parent category and all its children
        try:
            parent_category = Category.objects.get(id=parent_category_id)
            child_categories = Category.objects.filter(parent=parent_category)
            # Filter by parent category or any of its children
            category_ids = [parent_category.id] + list(child_categories.values_list('id', flat=True))
            products = products.filter(category_id__in=category_ids)
        except Category.DoesNotExist:
            pass
    
    if child_category_id:
        # Override parent category filter with specific child category
        products = products.filter(category_id=child_category_id)
    
    if brand_id:
        products = products.filter(brand_id=brand_id)
    if min_price:
        products = products.filter(base_price__gte=float(min_price))
    if max_price:
        products = products.filter(base_price__lte=float(max_price))
    if condition and condition != 'any':
        products = products.filter(condition=condition)
    if rating:
        products = products.filter(rating_avg__gte=float(rating))
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(category__name__icontains=search_query) |
            Q(brand__name__icontains=search_query)
        )
    if verified_only:
        products = products.filter(seller__is_verified=True)
    
            # Apply attribute filters with OR logic
        attribute_filters = Q()
        has_attribute_filters = False
        
        # Group attribute filters by attribute ID to handle multiple values per attribute
        attribute_groups = {}
        for key, value in request.GET.items():
            if key.startswith('attr_') and value:
                has_attribute_filters = True
                attr_id = key.replace('attr_', '')
                if attr_id not in attribute_groups:
                    attribute_groups[attr_id] = []
                attribute_groups[attr_id].append(value)
        
        if has_attribute_filters:
            # For each attribute, create OR condition for its values
            for attr_id, values in attribute_groups.items():
                attr_filter = Q()
                for value in values:
                    attr_filter |= Q(attribute_values__attribute_id=attr_id, attribute_values__value=value)
                # Combine all attributes with OR logic (product must match at least one of the selected attributes)
                attribute_filters |= attr_filter
            
            products = products.filter(attribute_filters).distinct()
    
    # Apply sorting
    if sort_by == 'price_low':
        products = products.order_by('base_price')
    elif sort_by == 'price_high':
        products = products.order_by('-base_price')
    elif sort_by == 'newest':
        products = products.order_by('-created_at')
    elif sort_by == 'rating':
        products = products.order_by('-rating_avg')
    elif sort_by == 'orders':
        products = products.order_by('-order_count')
    else:  # featured
        products = products.order_by('-order_count', '-rating_avg')
    
    # Pagination
    paginator = Paginator(products, 12)  # 12 products per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options - only parent categories initially
    parent_categories = Category.objects.filter(parent__isnull=True)
    brands = Brand.objects.all()
    
    # Calculate price range for slider
    if products.exists():
        min_price_range = products.aggregate(Min('base_price'))['base_price__min']
        max_price_range = products.aggregate(Max('base_price'))['base_price__max']
    else:
        min_price_range = 0
        max_price_range = 1000
    
    context = {
        'page_obj': page_obj,
        'parent_categories': parent_categories,
        'brands': brands,
        'min_price_range': min_price_range,
        'max_price_range': max_price_range,
        'current_filters': {
            'parent_category': parent_category_id,
            'child_category': child_category_id,
            'brand': brand_id,
            'min_price': min_price,
            'max_price': max_price,
            'condition': condition,
            'rating': rating,
            'search': search_query,
            'sort': sort_by,
            'view': view_type,
            'verified': verified_only,
        },
        'total_products': products.count(),
        'view_type': view_type,
    }
    
    return render(request, 'seller/ecom_product_listing.html', context)

@csrf_exempt
def get_filtered_products(request):
    """AJAX endpoint to get filtered products without page reload"""
    try:
        # Get all products with related data
        products = Product.objects.select_related('seller', 'category', 'brand').prefetch_related('images').all()
        
        # Get filter parameters
        parent_category_id = request.GET.get('parent_category')
        child_category_id = request.GET.get('child_category')
        brand_id = request.GET.get('brand')
        min_price = request.GET.get('min_price')
        max_price = request.GET.get('max_price')
        condition = request.GET.get('condition')
        rating = request.GET.get('rating')
        search_query = request.GET.get('search')
        sort_by = request.GET.get('sort', 'featured')
        verified_only = request.GET.get('verified') == 'on'
        
        # Apply filters
        if parent_category_id:
            # Get the parent category and all its children
            try:
                parent_category = Category.objects.get(id=parent_category_id)
                child_categories = Category.objects.filter(parent=parent_category)
                # Filter by parent category or any of its children
                category_ids = [parent_category.id] + list(child_categories.values_list('id', flat=True))
                products = products.filter(category_id__in=category_ids)
            except Category.DoesNotExist:
                pass
        
        if child_category_id:
            # Override parent category filter with specific child category
            products = products.filter(category_id=child_category_id)
        
        if brand_id:
            products = products.filter(brand_id=brand_id)
        if min_price:
            products = products.filter(base_price__gte=float(min_price))
        if max_price:
            products = products.filter(base_price__lte=float(max_price))
        if condition and condition != 'any':
            products = products.filter(condition=condition)
        if rating:
            products = products.filter(rating_avg__gte=float(rating))
        if search_query:
            products = products.filter(
                Q(name__icontains=search_query) | 
                Q(description__icontains=search_query) |
                Q(category__name__icontains=search_query) |
                Q(brand__name__icontains=search_query)
            )
        if verified_only:
            products = products.filter(seller__is_verified=True)
        
        # Apply attribute filters with OR logic
        attribute_filters = Q()
        has_attribute_filters = False
        
        # Group attribute filters by attribute ID to handle multiple values per attribute
        attribute_groups = {}
        for key, value in request.GET.items():
            if key.startswith('attr_') and value:
                has_attribute_filters = True
                attr_id = key.replace('attr_', '')
                if attr_id not in attribute_groups:
                    attribute_groups[attr_id] = []
                attribute_groups[attr_id].append(value)
        
        if has_attribute_filters:
            # For each attribute, create OR condition for its values
            for attr_id, values in attribute_groups.items():
                attr_filter = Q()
                for value in values:
                    attr_filter |= Q(attribute_values__attribute_id=attr_id, attribute_values__value=value)
                # Combine all attributes with OR logic (product must match at least one of the selected attributes)
                attribute_filters |= attr_filter
            
            products = products.filter(attribute_filters).distinct()
        
        # Apply sorting
        if sort_by == 'price_low':
            products = products.order_by('base_price')
        elif sort_by == 'price_high':
            products = products.order_by('-base_price')
        elif sort_by == 'newest':
            products = products.order_by('-created_at')
        elif sort_by == 'rating':
            products = products.order_by('-rating_avg')
        elif sort_by == 'orders':
            products = products.order_by('-order_count')
        else:  # featured
            products = products.order_by('-order_count', '-rating_avg')
        
        # Pagination
        paginator = Paginator(products, 12)  # 12 products per page
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Prepare products data for JSON response
        products_data = []
        for product in page_obj:
            # Get primary image
            primary_image = product.images.filter(is_thumbnail=True).first()
            if not primary_image:
                primary_image = product.images.first()
            
            # Convert image to base64
            image_url = ''
            if primary_image and primary_image.image:
                import base64
                image_url = f"data:image/jpeg;base64,{base64.b64encode(primary_image.image).decode('utf-8')}"
            
            # Get all images for the product
            all_images = product.images.all()
            images_data = []
            for img in all_images:
                if img.image:
                    img_base64 = base64.b64encode(img.image).decode('utf-8')
                    images_data.append({
                        'url': f"data:image/jpeg;base64,{img_base64}",
                        'is_thumbnail': img.is_thumbnail
                    })
            
            # Calculate current price (base_price - discount)
            current_price = product.base_price
            if product.discount_percentage:
                current_price = product.base_price * (1 - product.discount_percentage / 100)
            elif product.final_price:
                current_price = product.final_price
            
            product_data = {
                'id': product.id,
                'name': product.name,
                'description': product.description,
                'base_price': float(product.base_price),
                'current_price': float(current_price),
                'original_price': float(product.base_price) if product.discount_percentage or product.final_price else None,
                'rating_avg': float(product.rating_avg) if product.rating_avg else 0,
                'rating_count': product.reviews.count(),
                'order_count': product.order_count,
                'condition': product.condition,
                'stock_quantity': product.stock,
                'category_name': product.category.name,
                'brand_name': product.brand.name if product.brand else '',
                'seller_name': product.seller.user.username if product.seller.user else product.seller.name,
                'seller_verified': getattr(product.seller, 'is_verified', False),
                'primary_image': image_url,
                'images': images_data,
                'created_at': product.created_at.isoformat(),
            }
            products_data.append(product_data)
        
        response_data = {
            'success': True,
            'products': products_data,
            'total_products': products.count(),
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'current_page': page_obj.number,
            'total_pages': page_obj.paginator.num_pages,
            'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def get_buyer_info(request, buyer_id):
    """Get buyer information for seller dashboard"""
    seller = get_object_or_404(Seller, user=request.user)
    
    try:
        from buyer.models import Buyer
        buyer = Buyer.objects.get(id=buyer_id)
        
        # Get buyer's orders from this seller
        from buyer.models import Order
        orders = Order.objects.filter(buyer=buyer, items__product__seller=seller).distinct()
        
        return JsonResponse({
            'success': True,
            'buyer': {
            'id': buyer.id,
            'name': buyer.name,
            'email': buyer.email,
                'phone': buyer.phone,
                'address': buyer.address,
                'total_orders': orders.count(),
                'total_spent': sum(order.total for order in orders),
                'last_order_date': orders.order_by('-created_at').first().created_at if orders.exists() else None
            }
        })
    except Buyer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Buyer not found'})

# ========== REVIEW VIEWS ==========

@login_required
def submit_product_review(request, product_id):
    """Submit a product review"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get buyer
        from buyer.models import Buyer
        buyer = Buyer.objects.get(email=request.user.email)
        
        # Get product
        product = get_object_or_404(Product, id=product_id)
        
        # Check if buyer already reviewed this product
        existing_review = ProductReview.objects.filter(buyer=buyer, product=product).first()
        
        if existing_review:
            return JsonResponse({'success': False, 'error': 'You have already reviewed this product'})
        
        # Create review
        rating = float(request.POST.get('rating', 0))
        comment = request.POST.get('comment', '').strip()
        
        if not comment:
            return JsonResponse({'success': False, 'error': 'Comment is required'})
        
        if rating < 0 or rating > 5:
            return JsonResponse({'success': False, 'error': 'Rating must be between 0 and 5'})
        
        review = ProductReview.objects.create(
            buyer=buyer,
            product=product,
            rating=rating,
            comment=comment
        )
        
        # Update product's average rating
        product.update_rating_avg()
        
        return JsonResponse({
            'success': True,
            'message': 'Review submitted successfully',
            'review': {
                'id': review.id,
                'rating': review.rating,
                'comment': review.comment,
                'buyer_name': review.buyer.name,
                'created_at': review.created_at.strftime('%B %d, %Y'),
                'likes_count': 0,
                'dislikes_count': 0
            }
        })
        
    except Buyer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Buyer not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def submit_seller_review(request, seller_id):
    """Submit a seller review"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get buyer
        from buyer.models import Buyer
        buyer = Buyer.objects.get(email=request.user.email)
        
        # Get seller
        seller = get_object_or_404(Seller, id=seller_id)
        
        # Check if buyer already reviewed this seller
        existing_review = SellerReview.objects.filter(buyer=buyer, seller=seller).first()
        
        if existing_review:
            return JsonResponse({'success': False, 'error': 'You have already reviewed this seller'})
        
        # Get form data
        rating = request.POST.get('rating')
        comment = request.POST.get('comment', '').strip()
        
        if not rating:
            return JsonResponse({'success': False, 'error': 'Rating is required'})
        
        if not comment:
            return JsonResponse({'success': False, 'error': 'Comment is required'})
        
        # Validate rating
        try:
            rating = float(rating)
            if rating < 0.0 or rating > 5.0:
                raise ValueError("Rating must be between 0.0 and 5.0")
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid rating value'})
        
        review = SellerReview.objects.create(
            buyer=buyer,
            seller=seller,
            rating=rating,
            comment=comment
        )
        
        # Update seller's average rating
        seller.update_rating_avg()
        
        return JsonResponse({
            'success': True,
            'message': 'Review submitted successfully',
            'review': {
                'id': review.id,
                'comment': review.comment,
                'buyer_name': review.buyer.name,
                'created_at': review.created_at.strftime('%B %d, %Y'),
                'likes_count': 0,
                'dislikes_count': 0
            }
        })
        
    except Buyer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Buyer not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def like_review(request, review_type, review_id):
    """Like or dislike a review"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        # Get buyer
        from buyer.models import Buyer
        buyer = Buyer.objects.get(email=request.user.email)
        
        is_like = request.POST.get('is_like', 'true').lower() == 'true'
        
        if review_type == 'product':
            review = get_object_or_404(ProductReview, id=review_id)
            like_model = ProductReviewLike
        elif review_type == 'seller':
            review = get_object_or_404(SellerReview, id=review_id)
            like_model = SellerReviewLike
        else:
            return JsonResponse({'success': False, 'error': 'Invalid review type'})
        
        # Check if buyer already liked/disliked this review
        existing_like = like_model.objects.filter(buyer=buyer, review=review).first()
        
        if existing_like:
            if existing_like.is_like == is_like:
                # Remove like/dislike
                existing_like.delete()
                action = 'removed'
            else:
                # Change like/dislike
                existing_like.is_like = is_like
                existing_like.save()
                action = 'changed'
        else:
            # Create new like/dislike
            like_model.objects.create(buyer=buyer, review=review, is_like=is_like)
            action = 'added'
        
        return JsonResponse({
            'success': True,
            'message': f'{action.title()} successfully',
            'likes_count': review.likes_count,
            'dislikes_count': review.dislikes_count
        })
        
    except Buyer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Buyer not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_product_reviews(request, product_id):
    """Get product reviews for AJAX loading"""
    try:
        product = get_object_or_404(Product, id=product_id)
        reviews = ProductReview.objects.filter(product=product).select_related('buyer')
        
        reviews_data = []
        for review in reviews:
            reviews_data.append({
                'id': review.id,
                'rating': review.rating,
                'comment': review.comment,
                'buyer_name': review.buyer.name,
                'created_at': review.created_at.strftime('%B %d, %Y'),
                'likes_count': review.likes_count,
                'dislikes_count': review.dislikes_count
            })
        
        return JsonResponse({
            'success': True,
            'reviews': reviews_data,
            'average_rating': product.rating_avg or 0,
            'total_reviews': reviews.count()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def get_seller_reviews(request, seller_id):
    """Get seller reviews for AJAX loading"""
    try:
        seller = get_object_or_404(Seller, id=seller_id)
        reviews = SellerReview.objects.filter(seller=seller).select_related('buyer')
        
        reviews_data = []
        for review in reviews:
            reviews_data.append({
                'id': review.id,
                'rating': review.rating,
                'comment': review.comment,
                'buyer_name': review.buyer.name,
                'created_at': review.created_at.strftime('%B %d, %Y'),
                'likes_count': review.likes_count,
                'dislikes_count': review.dislikes_count
            })
        
        return JsonResponse({
            'success': True,
            'reviews': reviews_data,
            'average_rating': seller.rating or 0,
            'total_reviews': reviews.count()
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def about_view(request):
    """About Us page view"""
    context = {
        'title': 'About Us - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/about.html', context)


def find_store_view(request):
    """Find Store page view"""
    context = {
        'title': 'Find Store - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/find_store.html', context)


def partnership_view(request):
    """Partnership page view"""
    context = {
        'title': 'Partnership - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/partnership.html', context)


def information_view(request):
    """Information page view"""
    context = {
        'title': 'Information - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/information.html', context)


def money_refund_view(request):
    """Money Refund Policy page view"""
    context = {
        'title': 'Money Refund Policy - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/money_refund.html', context)


def shipping_view(request):
    """Shipping Policy page view"""
    context = {
        'title': 'Shipping Policy - MarketVibe',
        'user': request.user if request.user.is_authenticated else None,
    }
    return render(request, 'seller/shipping.html', context)
